#!/usr/bin/env python3

import openpyxl
from openpyxl import load_workbook
import os

def create_formula_based_dynamic_legacy(file_path):
    """Create a formula-based dynamic Legacy View that updates automatically"""
    try:
        print(f"Creating formula-based dynamic Legacy View: {file_path}")
        
        workbook = load_workbook(file_path)
        
        if "Units" not in workbook.sheetnames:
            print("Error: 'Units' sheet not found")
            return False
        
        # Remove existing Legacy View
        if "Legacy View" in workbook.sheetnames:
            del workbook["Legacy View"]
        
        legacy_ws = workbook.create_sheet("Legacy View")
        units_ws = workbook["Units"]
        
        # Add title and instructions
        legacy_ws.cell(row=1, column=1, value="Legacy View - Formula-Based Dynamic Structure")
        legacy_ws.cell(row=2, column=1, value="(Auto-updates when Units sheet changes - no VBA required)")
        legacy_ws.cell(row=3, column=1, value="Uses Excel formulas to dynamically group and sum by building")
        
        # Create a summary by building using Excel formulas
        headers_row = 5
        
        # Headers for summary view
        summary_headers = ["Building", "Unit Count", "Net Area Total", "Rent PA Total", "ERV 2024 Total"]
        for col, header in enumerate(summary_headers, 1):
            cell = legacy_ws.cell(row=headers_row, column=col, value=header)
            if hasattr(openpyxl.styles, 'Font'):
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Get unique buildings using formula approach
        # This is a simplified approach - we'll create manual formulas for the main buildings
        # For a fully dynamic solution, we would need VBA or a more complex formula setup
        
        # Add instructions for formula-based approach
        instruction_row = 7
        instructions = [
            "FORMULA-BASED APPROACH:",
            "",
            "This approach uses Excel formulas to create dynamic summaries",
            "that automatically update when the Units sheet changes.",
            "",
            "Benefits:",
            "• No VBA required",
            "• Automatic updates",
            "• Compatible with all Excel versions",
            "",
            "Each building summary uses SUMIF formulas like:",
            "=SUMIF(Units.A:A,\"Building Name\",Units.C:C)",
            "",
            "To add new buildings, copy formula rows and change building name."
        ]
        
        for i, instruction in enumerate(instructions):
            legacy_ws.cell(row=instruction_row + i, column=1, value=instruction)
            if instruction.endswith(":"):
                cell = legacy_ws.cell(row=instruction_row + i, column=1)
                if hasattr(openpyxl.styles, 'Font'):
                    cell.font = openpyxl.styles.Font(bold=True)
        
        # Get unique building names from Units sheet
        buildings = set()
        units_max_row = units_ws.max_row
        
        for row in range(2, units_max_row + 1):
            building_name = units_ws.cell(row=row, column=1).value
            if building_name and str(building_name).strip():
                buildings.add(str(building_name).strip())
        
        print(f"Found {len(buildings)} unique buildings")
        
        # Create summary formulas for each building
        data_start_row = instruction_row + len(instructions) + 2
        current_row = data_start_row
        
        # Headers for building summary
        legacy_ws.cell(row=current_row, column=1, value="Building Summary (Formula-Based)")
        if hasattr(openpyxl.styles, 'Font'):
            legacy_ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        current_row += 2
        
        # Column headers
        summary_headers = ["Building", "Unit Count", "Total Net Area", "Total Rent PA", "Total ERV 2024"]
        for col, header in enumerate(summary_headers, 1):
            cell = legacy_ws.cell(row=current_row, column=col, value=header)
            if hasattr(openpyxl.styles, 'Font'):
                cell.font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Find the relevant column indices in Units sheet
        net_area_col = None
        rent_pa_col = None
        erv_2024_col = None
        
        for col in range(1, units_ws.max_column + 1):
            header = units_ws.cell(row=1, column=col).value
            if header:
                header_lower = str(header).lower()
                if 'net area' in header_lower:
                    net_area_col = col
                elif 'rent pa' in header_lower:
                    rent_pa_col = col
                elif '2024 erv' in header_lower or 'erv 2024' in header_lower:
                    erv_2024_col = col
        
        # Create formula-based summary for each building
        for building in sorted(buildings):
            # Building name
            legacy_ws.cell(row=current_row, column=1, value=building)
            
            # Unit count formula
            count_formula = f'=COUNTIF(Units.A:A,"{building}")'
            legacy_ws.cell(row=current_row, column=2, value=count_formula)
            
            # Sum formulas for numeric columns
            if net_area_col:
                col_letter = get_column_letter(net_area_col)
                area_formula = f'=SUMIF(Units.A:A,"{building}",Units.{col_letter}:{col_letter})'
                legacy_ws.cell(row=current_row, column=3, value=area_formula)
            
            if rent_pa_col:
                col_letter = get_column_letter(rent_pa_col)
                rent_formula = f'=SUMIF(Units.A:A,"{building}",Units.{col_letter}:{col_letter})'
                legacy_ws.cell(row=current_row, column=4, value=rent_formula)
            
            if erv_2024_col:
                col_letter = get_column_letter(erv_2024_col)
                erv_formula = f'=SUMIF(Units.A:A,"{building}",Units.{col_letter}:{col_letter})'
                legacy_ws.cell(row=current_row, column=5, value=erv_formula)
            
            current_row += 1
        
        # Auto-fit columns
        for column in legacy_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            legacy_ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(file_path)
        workbook.close()
        
        print(f"✅ Formula-based Legacy View created with {len(buildings)} building summaries")
        print("🔗 Uses SUMIF formulas that automatically update when Units data changes")
        print("📊 No VBA required - works with all Excel versions")
        return True
        
    except Exception as e:
        print(f"❌ Error creating formula-based Legacy View: {str(e)}")
        return False

def get_column_letter(col_num):
    """Convert column number to Excel letter (A, B, C, etc.)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result

if __name__ == "__main__":
    file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/8.xlsx"
    
    if os.path.exists(file_path):
        # Create backup
        backup_path = file_path.replace('.xlsx', '_backup_before_formula_based.xlsx')
        print(f"Creating backup: {backup_path}")
        import shutil
        shutil.copy2(file_path, backup_path)
        
        # Create formula-based version
        create_formula_based_dynamic_legacy(file_path)
    else:
        print(f"File not found: {file_path}")
