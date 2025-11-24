#!/usr/bin/env python3
"""
Enhanced Bank Schedule Sanitizer with Dynamic Legacy View
"""
import os
import sys
sys.path.append('.')

import openpyxl
import tkinter as tk
from bank_schedule_sanitizer import BankScheduleSanitizer

class DynamicLegacyBankScheduleSanitizer(BankScheduleSanitizer):
    """Enhanced version with dynamic Legacy View that auto-updates from Units sheet"""
    
    def create_hierarchical_view_sheet(self, workbook, units_ws, units_row):
        """Create a dynamic Legacy View sheet with formulas that auto-update from Units sheet"""
        try:
            self.log_message("📋 Creating dynamic Legacy View sheet...")
            
            # Remove any existing legacy view sheets
            sheets_to_remove = ["Hierarchical View", "Hierarchical View1", "Legacy View"]
            for sheet_name in sheets_to_remove:
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]
                
            hier_ws = workbook.create_sheet("Legacy View")
            
            # Add title and instructions
            hier_ws.cell(row=1, column=1, value="Legacy View - Dynamic Building Structure")
            hier_ws.cell(row=2, column=1, value="(Auto-updates from Units sheet - formulas link to live data)")
            
            # Set up headers in row 4
            headers_row = 4
            
            # Create column mapping and copy headers
            col_mapping = {}
            target_col = 1
            
            for source_col in range(1, units_ws.max_column + 1):
                header = units_ws.cell(row=1, column=source_col).value
                if header and str(header).strip():
                    header_str = str(header).strip()
                    col_mapping[source_col] = target_col
                    hier_ws.cell(row=headers_row, column=target_col, value=header_str)
                    target_col += 1
            
            # Create formulas that reference Units sheet data
            legacy_row = headers_row + 1
            
            for units_data_row in range(2, units_row):
                for source_col, target_col in col_mapping.items():
                    # Create formula that dynamically references Units sheet
                    col_letter = self.get_column_letter(source_col)
                    formula = f"=Units.{col_letter}{units_data_row}"
                    hier_ws.cell(row=legacy_row, column=target_col, value=formula)
                
                legacy_row += 1
            
            # Format headers
            for col in range(1, len(col_mapping) + 1):
                cell = hier_ws.cell(row=headers_row, column=col)
                if hasattr(openpyxl.styles, 'Font'):
                    cell.font = openpyxl.styles.Font(bold=True)
                
            # Auto-width columns
            try:
                for column in hier_ws.columns:
                    max_length = 0
                    column_letter = self.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    hier_ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                self.log_message(f"⚠️ Column auto-width failed: {e}")
            
            self.log_message("✅ Legacy View sheet created with dynamic formulas")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating dynamic Legacy View sheet: {str(e)}")
            return False
    
    def get_column_letter(self, col_num):
        """Convert column number to Excel letter (A, B, C, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

def test_dynamic_legacy_view():
    """Test the dynamic Legacy View functionality"""
    # Create root window but don't show it
    root = tk.Tk()
    root.withdraw()  # Hide the window
    
    # Create enhanced sanitizer
    sanitizer = DynamicLegacyBankScheduleSanitizer(root)
    
    # Input and output paths
    input_path = 'data/Leasing Bank Schedule June 2025.xlsx'
    output_path = 'dynamic_legacy_view_test.xlsx'
    
    print("🧪 Testing Dynamic Legacy View implementation...")
    
    try:
        # Analyze the data first
        analysis_results = sanitizer.analyze_bank_schedule_data(input_path)
        print(f"📊 Analysis: {analysis_results['buildings']} buildings, {analysis_results['units']} units")
        
        # Create the output file with the individual methods
        sanitizer.create_buildings_summary_sheet(input_path, output_path, analysis_results)
        sanitizer.create_units_sheet(input_path, output_path, analysis_results['building_names'])
        sanitizer.create_hierarchical_view_sheet_from_files(input_path, output_path, analysis_results['building_names'])
        
        # Check the results
        if os.path.exists(output_path):
            wb = openpyxl.load_workbook(output_path)
            print(f"📋 Created sheets: {wb.sheetnames}")
            
            # Check for Legacy View sheets
            legacy_sheets = [s for s in wb.sheetnames if 'Legacy' in s]
            print(f"🏛️ Legacy View sheets: {legacy_sheets}")
            
            if len(legacy_sheets) == 1 and legacy_sheets[0] == 'Legacy View':
                print("✅ SUCCESS: Single 'Legacy View' sheet created!")
                
                # Check the structure
                ws = wb['Legacy View']
                print(f"✅ Legacy View structure: {ws.max_column} columns × {ws.max_row} rows")
                
                # Check for formulas (look for cells starting with =)
                formula_count = 0
                sample_formulas = []
                for row in range(5, min(10, ws.max_row + 1)):
                    for col in range(1, min(6, ws.max_column + 1)):
                        cell = ws.cell(row, col)
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1
                            if len(sample_formulas) < 3:
                                sample_formulas.append(f"Cell {chr(64+col)}{row}: {cell.value}")
                
                print(f"🔗 Found {formula_count} formulas linking to Units sheet")
                if sample_formulas:
                    print("📝 Sample formulas:")
                    for formula in sample_formulas:
                        print(f"   {formula}")
                
            else:
                print(f"❌ ISSUE: Expected single 'Legacy View' sheet, got: {legacy_sheets}")
            
            wb.close()
            print(f"📁 Output saved to: {output_path}")
            
        else:
            print("❌ ERROR: Output file not created")
            
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    # Clean up
    root.destroy()

if __name__ == "__main__":
    test_dynamic_legacy_view()
