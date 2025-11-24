#!/usr/bin/env python3
"""
Test the improved Legacy View functionality
"""

import pandas as pd
import sys
import os
import tempfile
import tkinter as tk

# Add the current directory to path so we can import our main module
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from bank_schedule_sanitizer import BankScheduleSanitizer

def test_legacy_view():
    """Test the Legacy View creation with proper column structure."""
    
    print("Testing Improved Legacy View...")
    print("="*50)
    
    test_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/Leasing Bank Schedule June 2025.xlsx"
    
    if not os.path.exists(test_file):
        print(f"❌ Test file not found: {test_file}")
        return False
    
    try:
        # Create a minimal tkinter instance (needed for the class)
        root = tk.Tk()
        root.withdraw()  # Hide the window for testing
        
        # Create an instance of our sanitizer
        sanitizer = BankScheduleSanitizer(root)
        
        # Override the log_message method to print to console for testing
        def test_log(message):
            print(f"[LOG] {message}")
        sanitizer.log_message = test_log
        
        # Create a temporary output file for testing
        output_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/test_legacy_view.xlsx"
        
        print(f"📊 Running Legacy View test...")
        print(f"📄 Input:  {os.path.basename(test_file)}")
        print(f"📁 Output: {os.path.basename(output_path)}")
        
        # Run analysis
        analysis_results = sanitizer.analyze_bank_schedule_data(test_file)
        
        # Create Buildings sheet
        print(f"📊 Creating Buildings sheet...")
        sanitizer.create_buildings_summary_sheet(test_file, output_path, analysis_results)
        
        # Create Units sheet  
        print(f"📋 Creating Units sheet...")
        sanitizer.create_units_sheet(test_file, output_path, analysis_results['building_names'])
        
        # Create Legacy View sheet
        print(f"🏛️ Creating Legacy View sheet...")
        success = sanitizer.create_hierarchical_view_sheet_from_files(test_file, output_path, analysis_results['building_names'])
        
        if success:
            # Verify the output
            import openpyxl
            workbook = openpyxl.load_workbook(output_path)
            sheets = workbook.sheetnames
            
            print(f"✅ Success! Sheets created: {sheets}")
            
            # Check the Legacy View sheet specifically
            if 'Legacy View' in sheets:
                legacy_ws = workbook['Legacy View']
                rows, cols = legacy_ws.max_row, legacy_ws.max_column
                print(f"🏛️ Legacy View: {rows} rows × {cols} columns")
                
                # Check headers
                print(f"📋 Headers (first 10):")
                for col in range(1, min(11, cols + 1)):
                    header = legacy_ws.cell(row=4, column=col).value
                    if header:
                        print(f"   Col {col}: {header}")
                
                # Check structure
                building_count = 0
                unit_count = 0
                
                for row in range(5, min(20, rows + 1)):
                    cell_value = legacy_ws.cell(row=row, column=1).value
                    if cell_value:
                        cell_str = str(cell_value)
                        if not cell_str.startswith("  "):  # Building (not indented)
                            building_count += 1
                            print(f"🏢 Building: {cell_str}")
                        elif cell_str.startswith("  "):  # Unit (indented)
                            unit_count += 1
                            if unit_count <= 3:  # Show first few units
                                print(f"   📦 Unit: {cell_str.strip()}")
                
                print(f"✅ Found {building_count} buildings and {unit_count} units")
                
                # Check that we don't have duplicate sheets
                hierarchical_sheets = [s for s in sheets if 'hierarchical' in s.lower() or 'legacy' in s.lower()]
                print(f"🔍 Hierarchical-type sheets: {hierarchical_sheets}")
                
            workbook.close()
            print(f"\n🎉 Legacy View test completed successfully!")
            return True
            
        else:
            print(f"❌ Legacy View creation failed")
            return False
        
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_legacy_view()
    if not success:
        sys.exit(1)
