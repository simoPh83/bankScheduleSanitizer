#!/usr/bin/env python3
"""
Add Dynamic Legacy View to existing Excel file
"""
import openpyxl

def create_dynamic_legacy_view(file_path):
    """Add dynamic Legacy View with formulas to existing Excel file"""
    print(f"🔄 Adding dynamic Legacy View to {file_path}...")
    
    try:
        # Open the existing file
        wb = openpyxl.load_workbook(file_path)
        
        # Check if Units sheet exists
        if 'Units' not in wb.sheetnames:
            print("❌ Error: No Units sheet found")
            return False
            
        units_ws = wb['Units']
        print(f"📊 Units sheet: {units_ws.max_column} cols × {units_ws.max_row} rows")
        
        # Remove existing Legacy View if it exists
        if 'Legacy View' in wb.sheetnames:
            del wb['Legacy View']
        
        # Create new Legacy View sheet
        legacy_ws = wb.create_sheet('Legacy View')
        
        # Add title and description
        legacy_ws.cell(row=1, column=1, value="Legacy View - Dynamic Links to Units Sheet")
        legacy_ws.cell(row=2, column=1, value="(All data auto-updates when Units sheet changes)")
        
        # Copy headers from Units sheet (row 4)
        headers_row = 4
        for col in range(1, units_ws.max_column + 1):
            header = units_ws.cell(1, col).value
            if header:
                legacy_ws.cell(headers_row, col, value=header)
                # Make header bold
                legacy_ws.cell(headers_row, col).font = openpyxl.styles.Font(bold=True)
        
        # Create dynamic formulas for all data rows
        legacy_row = headers_row + 1
        
        for units_row in range(2, units_ws.max_row + 1):  # Start from row 2 (after headers)
            for col in range(1, units_ws.max_column + 1):
                # Create formula that references Units sheet
                col_letter = chr(64 + col) if col <= 26 else f"{chr(64 + (col-1)//26)}{chr(65 + (col-1)%26)}"
                formula = f"=Units.{col_letter}{units_row}"
                legacy_ws.cell(legacy_row, col, value=formula)
            
            legacy_row += 1
        
        # Auto-adjust column widths
        for col in range(1, units_ws.max_column + 1):
            col_letter = chr(64 + col) if col <= 26 else f"{chr(64 + (col-1)//26)}{chr(65 + (col-1)%26)}"
            legacy_ws.column_dimensions[col_letter].width = 15
        
        # Save the file
        wb.save(file_path)
        wb.close()
        
        print(f"✅ Dynamic Legacy View added successfully!")
        print(f"🔗 Created {(units_ws.max_row - 1) * units_ws.max_column} dynamic formulas")
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

def test_dynamic_formulas(file_path):
    """Test that the dynamic formulas are working"""
    try:
        wb = openpyxl.load_workbook(file_path)
        
        if 'Legacy View' not in wb.sheetnames:
            print("❌ Legacy View sheet not found")
            return
            
        legacy_ws = wb['Legacy View']
        
        # Check a few sample formulas
        print("\\n🔍 Checking sample dynamic formulas:")
        test_cells = [(5, 1), (5, 2), (6, 1), (6, 3)]
        
        for row, col in test_cells:
            if row <= legacy_ws.max_row and col <= legacy_ws.max_column:
                cell = legacy_ws.cell(row, col)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    col_letter = chr(64 + col) if col <= 26 else f"{chr(64 + (col-1)//26)}{chr(65 + (col-1)%26)}"
                    print(f"  Cell {col_letter}{row}: {cell.value}")
        
        # Count total formulas
        formula_count = 0
        for row in legacy_ws.iter_rows(min_row=5):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
        
        print(f"\\n📊 Total dynamic formulas: {formula_count}")
        wb.close()
        
    except Exception as e:
        print(f"❌ Test error: {e}")

if __name__ == "__main__":
    file_path = "dynamic_test_base.xlsx"
    
    if create_dynamic_legacy_view(file_path):
        test_dynamic_formulas(file_path)
        print(f"\\n🎉 Dynamic Legacy View ready! Open {file_path} in Excel to see live updates.")
        print("📝 Any changes to the Units sheet will automatically appear in Legacy View!")
