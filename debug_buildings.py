#!/usr/bin/env python3
"""
Quick debug script to check the Buildings sheet content
"""

import openpyxl
import os

# Find the latest numbered file
data_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data"
files = [f for f in os.listdir(data_path) if f.endswith('.xlsx') and f.split('.')[0].isdigit()]
if files:
    latest_file = max(files, key=lambda x: int(x.split('.')[0]))
    file_path = os.path.join(data_path, latest_file)
    
    print(f"Checking file: {latest_file}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        
        print(f"Sheets in workbook: {wb.sheetnames}")
        
        if "Buildings" in wb.sheetnames:
            buildings_ws = wb["Buildings"]
            print(f"\nBuildings sheet - Max row: {buildings_ws.max_row}, Max col: {buildings_ws.max_column}")
            
            # Check headers
            print("\nHeaders:")
            for col in range(1, buildings_ws.max_column + 1):
                header = buildings_ws.cell(row=1, column=col).value
                print(f"  Column {col}: {header}")
            
            # Check first few data rows
            print("\nFirst 3 data rows:")
            for row in range(2, min(5, buildings_ws.max_row + 1)):
                row_data = []
                for col in range(1, buildings_ws.max_column + 1):
                    cell = buildings_ws.cell(row=row, column=col)
                    row_data.append(str(cell.value)[:20] if cell.value else "")
                print(f"  Row {row}: {row_data}")
        
        if "Units" in wb.sheetnames:
            units_ws = wb["Units"]
            print(f"\nUnits sheet - Max row: {units_ws.max_row}, Max col: {units_ws.max_column}")
            print("Units headers:")
            for col in range(1, min(10, units_ws.max_column + 1)):
                header = units_ws.cell(row=1, column=col).value
                print(f"  Column {col}: {header}")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading file: {e}")
else:
    print("No numbered xlsx files found")
