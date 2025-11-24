#!/usr/bin/env python3
"""
Test Excel compatibility fixes
This script tests the bank schedule sanitizer with focus on Excel file compatibility
"""

import sys
import os
import subprocess
import shutil
from pathlib import Path

# Add the current directory to Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def test_excel_compatibility():
    """Test that the sanitizer creates Excel-compatible files"""
    print("Testing Excel compatibility fixes...")
    print("="*60)
    
    # Define paths
    input_file = "data/Leasing Bank Schedule June 2025.xlsx"
    output_file = "test_output_compatibility.xlsx"
    
    # Clean up any existing test output
    if os.path.exists(output_file):
        os.remove(output_file)
    
    try:
        # Import the sanitizer
        from bank_schedule_sanitizer import BankScheduleSanitizer
        import tkinter as tk
        
        # Create a test instance (headless)
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        
        app = BankScheduleSanitizer(root)
        
        # Set the input file path
        app.input_file_path.set(input_file)
        
        # Test the sanitization process
        print("Step 1: Testing file validation...")
        is_valid, message = app.validate_excel_file(input_file)
        print(f"   File validation: {'✅ PASS' if is_valid else '❌ FAIL'}")
        if not is_valid:
            print(f"   Error: {message}")
            return False
        
        print("\nStep 2: Testing data analysis...")
        analysis_results = app.analyze_bank_schedule_data(input_file)
        building_count = analysis_results['buildings']
        unit_count = analysis_results['units'] 
        empty_count = analysis_results['empty_rows']
        building_names = analysis_results['building_names']
        print(f"   Buildings: {building_count}")
        print(f"   Units: {unit_count}")
        print(f"   Empty rows: {empty_count}")
        print(f"   Building names count: {len(building_names)}")
        
        print("\nStep 3: Testing file copy and initial setup...")
        shutil.copy2(input_file, output_file)
        print(f"   ✅ File copied to {output_file}")
        
        print("\nStep 4: Testing Buildings sheet creation...")
        app.create_buildings_summary_sheet(input_file, output_file, analysis_results)
        print("   ✅ Buildings sheet created")
        
        print("\nStep 5: Testing Units sheet creation...")
        app.create_units_sheet(input_file, output_file, building_names)
        print("   ✅ Units sheet created")
        
        # Test file integrity
        print("\nStep 6: Testing Excel file integrity...")
        
        # Try to open with pandas to verify basic structure
        import pandas as pd
        try:
            # Test reading each sheet
            original_sheets = pd.ExcelFile(input_file).sheet_names
            output_sheets = pd.ExcelFile(output_file).sheet_names
            
            print(f"   Original sheets: {original_sheets}")
            print(f"   Output sheets: {output_sheets}")
            
            # Check that we have the expected new sheets
            expected_new_sheets = ['Buildings', 'Units']
            for sheet in expected_new_sheets:
                if sheet in output_sheets:
                    print(f"   ✅ {sheet} sheet exists")
                    
                    # Try to read the sheet
                    df = pd.read_excel(output_file, sheet_name=sheet, nrows=5)
                    print(f"      - {sheet} sheet has {len(df)} rows (sample)")
                    print(f"      - Columns: {list(df.columns)[:5]}...")
                else:
                    print(f"   ❌ {sheet} sheet missing")
                    return False
            
            print("   ✅ All sheets readable with pandas")
            
        except Exception as e:
            print(f"   ❌ Error reading Excel file with pandas: {str(e)}")
            return False
        
        # Test with openpyxl to check for XML issues
        print("\nStep 7: Testing with openpyxl (XML validation)...")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
            
            print(f"   Available worksheets: {wb.sheetnames}")
            
            # Test each sheet
            for sheet_name in ['Buildings', 'Units']:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"   ✅ {sheet_name} sheet - {ws.max_row} rows, {ws.max_column} columns")
                    
                    # Check for data validation
                    if hasattr(ws, 'data_validations') and ws.data_validations:
                        print(f"      - Data validations: {len(ws.data_validations.dataValidation)}")
                    else:
                        print(f"      - No data validations (safer for compatibility)")
                        
            wb.close()
            print("   ✅ openpyxl can read file without errors")
            
        except Exception as e:
            print(f"   ❌ Error with openpyxl: {str(e)}")
            return False
        
        # Try to simulate Excel opening (basic file structure check)
        print("\nStep 8: Basic file structure validation...")
        try:
            import zipfile
            with zipfile.ZipFile(output_file, 'r') as zip_file:
                file_list = zip_file.namelist()
                
                # Check for key Excel files
                key_files = ['xl/workbook.xml', 'xl/sharedStrings.xml', '[Content_Types].xml']
                for key_file in key_files:
                    if key_file in file_list:
                        print(f"   ✅ {key_file} present")
                    else:
                        print(f"   ⚠️ {key_file} missing")
                
                # Check worksheet files
                worksheet_files = [f for f in file_list if f.startswith('xl/worksheets/')]
                print(f"   ✅ Found {len(worksheet_files)} worksheet files")
                
        except Exception as e:
            print(f"   ⚠️ Could not validate ZIP structure: {str(e)}")
        
        print(f"\n✅ SUCCESS: Excel file created successfully at {output_file}")
        print("\nFile is ready for testing in Excel!")
        
        # File size check
        file_size = os.path.getsize(output_file)
        print(f"Output file size: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)")
        
        return True
        
    except Exception as e:
        print(f"\n❌ FAILED: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        try:
            root.destroy()
        except:
            pass

def main():
    """Main test function"""
    print("Bank Schedule Sanitizer - Excel Compatibility Test")
    print("="*60)
    
    # Check if input file exists
    input_file = "data/Leasing Bank Schedule June 2025.xlsx"
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        print("Make sure the data file is in the correct location.")
        return False
    
    success = test_excel_compatibility()
    
    if success:
        print("\n" + "="*60)
        print("🎉 ALL TESTS PASSED!")
        print("The Excel compatibility fixes appear to be working correctly.")
        print("\nNext steps:")
        print("1. Open the output file in Excel to verify it works")
        print("2. Verify all building names are populated in Units sheet")
        print("3. Test autofilter functionality for building search")
        print("4. Verify all formatting is preserved")
    else:
        print("\n" + "="*60)
        print("❌ SOME TESTS FAILED")
        print("Check the error messages above for details.")
    
    return success

if __name__ == "__main__":
    main()
