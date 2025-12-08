#!/usr/bin/env python3
"""
Status Column Updater

Updates the Status column (column AE) in Excel files by setting "Let" for units that meet criteria:
- Row contains unit data (non-empty row)
- Tenant Name is not vacant/null
- Start Date is present
- Expiry Date is present

Created: December 2024
"""

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import sys
from datetime import datetime

def is_file_locked(file_path):
    """
    Check if a file is locked/in use (especially for network drives on macOS)
    """
    try:
        # Try to open the file in append mode
        with open(file_path, 'r+b') as f:
            pass
        return False
    except (IOError, OSError, PermissionError):
        return True

def find_column_by_name(worksheet, column_name, search_rows=10):
    """
    Find a column by searching for its header name in the first few rows
    Returns the column letter (e.g., 'AE') or None if not found
    """
    for row in range(1, search_rows + 1):
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row, column=col).value
            if cell_value and isinstance(cell_value, str):
                # Clean up the cell value by removing extra spaces
                clean_cell_value = ' '.join(cell_value.split())
                clean_column_name = ' '.join(column_name.split())
                
                if clean_column_name.lower() in clean_cell_value.lower():
                    return openpyxl.utils.get_column_letter(col)
    return None

def is_vacant_tenant(tenant_name):
    """
    Check if tenant name indicates vacancy or is a non-tenant entry
    """
    if not tenant_name:
        return True
    
    tenant_str = str(tenant_name).strip().upper()
    
    # Common vacancy indicators
    vacancy_indicators = [
        'VACANT', 'VACANCY', 'EMPTY', 'AVAILABLE', 
        'TO LET', 'FOR RENT', 'UNOCCUPIED', 'NULL'
    ]
    
    # Update/system messages (not actual tenant names)
    system_messages = [
        'UPDATED', 'UPDATE', 'REVISED', 'CHANGED', 'MODIFIED',
        'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE',
        'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER',
        '2024', '2025', '2026', '2027', '2028', '2029', '2030'
    ]
    
    # Check if tenant name is just whitespace or common vacancy terms
    if not tenant_str or tenant_str in vacancy_indicators:
        return True
    
    # Check if tenant name contains vacancy indicators
    for indicator in vacancy_indicators:
        if indicator in tenant_str:
            return True
    
    # Check if this looks like a system message/update note
    for message in system_messages:
        if message in tenant_str:
            return True
    
    # Check if it looks like a date string (system update)
    if any(month in tenant_str for month in ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 
                                             'JULY', 'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']) \
       and any(year in tenant_str for year in ['2024', '2025', '2026', '2027', '2028', '2029']):
        return True
    
    return False

def has_date_value(cell_value):
    """
    Check if a cell contains a valid date value
    """
    if not cell_value:
        return False
    
    # Check if it's already a datetime object
    if isinstance(cell_value, datetime):
        return True
    
    # Try to parse as date string
    if isinstance(cell_value, str):
        cell_value = cell_value.strip()
        if not cell_value:
            return False
        
        # Common date formats to try
        date_formats = [
            '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
            '%Y/%m/%d', '%d.%m.%Y', '%Y.%m.%d'
        ]
        
        for date_format in date_formats:
            try:
                datetime.strptime(cell_value, date_format)
                return True
            except ValueError:
                continue
    
    # Check if it's a number (Excel date serial)
    try:
        float(cell_value)
        return True
    except (ValueError, TypeError):
        pass
    
    return False

def update_status_column(file_path, sheet_name=None, create_backup=True):
    """
    Update the Status column in the specified Excel file
    """
    print(f"\n=== Status Column Updater ===")
    print(f"Processing file: {os.path.basename(file_path)}")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"❌ Error: File not found: {file_path}")
        return False
    
    # Check if file is locked
    if is_file_locked(file_path):
        print(f"❌ Error: File is currently open in another application.")
        print(f"Please close the file and try again.")
        return False
    
    # Create backup if requested
    if create_backup:
        backup_path = file_path.replace('.xlsx', '_BACKUP_StatusUpdate.xlsx')
        try:
            import shutil
            shutil.copy2(file_path, backup_path)
            print(f"📋 Backup created: {os.path.basename(backup_path)}")
        except Exception as e:
            print(f"⚠️  Warning: Could not create backup: {e}")
    
    try:
        # Open workbook
        print("📂 Opening workbook...")
        workbook = openpyxl.load_workbook(file_path)
        
        # Determine which sheet to work with
        if sheet_name and sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"📋 Working with sheet: '{sheet_name}'")
        else:
            # Use the active sheet or first sheet
            worksheet = workbook.active
            print(f"📋 Working with sheet: '{worksheet.title}'")
        
        # Find key columns
        print("🔍 Locating columns...")
        
        # Find Status column (should be AE)
        status_col = find_column_by_name(worksheet, 'Status')
        if not status_col:
            # Fallback to column AE
            status_col = 'AE'
            print(f"⚠️  Status column not found by name, using column {status_col}")
        else:
            print(f"✅ Status column found: {status_col}")
        
        # Find other required columns
        tenant_col = find_column_by_name(worksheet, 'Tenant Name')
        
        # Try multiple variations for date columns
        start_date_col = (find_column_by_name(worksheet, 'Start Date') or 
                         find_column_by_name(worksheet, 'Lease Start') or
                         find_column_by_name(worksheet, 'Start') or
                         find_column_by_name(worksheet, 'Commencement'))
        
        expiry_date_col = (find_column_by_name(worksheet, 'Expiry Date') or
                          find_column_by_name(worksheet, 'End Date') or
                          find_column_by_name(worksheet, 'Lease End') or
                          find_column_by_name(worksheet, 'Expiry') or
                          find_column_by_name(worksheet, 'Termination'))
        
        print(f"📍 Column mapping:")
        print(f"   Status: {status_col}")
        print(f"   Tenant Name: {tenant_col or 'Not found'}")
        print(f"   Start Date: {start_date_col or 'Not found'}")
        print(f"   Expiry Date: {expiry_date_col or 'Not found'}")
        
        if not tenant_col or not start_date_col or not expiry_date_col:
            print("❌ Error: Could not find all required columns:")
            print(f"   Tenant Name: {'✓' if tenant_col else '✗'}")
            print(f"   Start Date: {'✓' if start_date_col else '✗'}")  
            print(f"   Expiry Date: {'✓' if expiry_date_col else '✗'}")
            print("   Cannot proceed without all required columns.")
            workbook.close()
            return False
        
        # Process rows
        print("\n🔄 Processing rows...")
        updates_made = 0
        rows_processed = 0
        
        # Start from row 3 to skip headers
        for row_num in range(3, worksheet.max_row + 1):
            rows_processed += 1
            
            # Check if this is a unit row (has some data in key columns)
            has_unit_data = False
            
            # Check key unit identification columns (Unit Demise, Unit Type, Net Area, etc.)
            unit_demise = worksheet.cell(row=row_num, column=6).value  # Column F
            unit_type = worksheet.cell(row=row_num, column=7).value    # Column G  
            net_area = worksheet.cell(row=row_num, column=8).value     # Column H
            
            # A proper unit row should have at least unit demise or unit type or net area
            if (unit_demise and str(unit_demise).strip()) or \
               (unit_type and str(unit_type).strip()) or \
               (net_area and str(net_area).strip()):
                has_unit_data = True

            # Get current status
            status_cell = worksheet[f"{status_col}{row_num}"]
            current_status = status_cell.value

            # If this row doesn't have unit data but has any status, remove it
            if not has_unit_data and current_status in ["LET", "Let"]:
                status_cell.value = None
                updates_made += 1
                print(f"   ❌ Row {row_num}: Removed '{current_status}' status (not a unit row)")
                continue

            # If this is not a unit row, skip further processing
            if not has_unit_data:
                continue

            # Check criteria for "Let" status - ALL must be true
            should_be_let = True
            reasons = []

            # 1. Check Tenant Name (not vacant)
            tenant_cell = worksheet[f"{tenant_col}{row_num}"]
            tenant_name = tenant_cell.value
            if is_vacant_tenant(tenant_name):
                should_be_let = False
                reasons.append("vacant tenant")

            # 2. Check Start Date (REQUIRED)
            start_date_cell = worksheet[f"{start_date_col}{row_num}"]
            if not has_date_value(start_date_cell.value):
                should_be_let = False
                reasons.append("no start date")

            # 3. Check Expiry Date (REQUIRED) 
            expiry_date_cell = worksheet[f"{expiry_date_col}{row_num}"]
            if not has_date_value(expiry_date_cell.value):
                should_be_let = False
                reasons.append("no expiry date")

            # Update status if needed
            if should_be_let and current_status not in ["LET", "Let"]:
                status_cell.value = "Let"
                updates_made += 1
                print(f"   ✅ Row {row_num}: Set status to 'Let'")
            elif should_be_let and current_status in ["LET", "Let"]:
                if current_status != "Let":
                    status_cell.value = "Let"
                    updates_made += 1
                    print(f"   🔄 Row {row_num}: Updated to 'Let' (was '{current_status}')")
                else:
                    print(f"   ✓  Row {row_num}: Already 'Let'")
            elif not should_be_let and current_status in ["LET", "Let"]:
                # Clear Let status if it doesn't meet criteria
                status_cell.value = None
                updates_made += 1
                print(f"   ❌ Row {row_num}: Removed '{current_status}' status ({', '.join(reasons)})")
            elif not should_be_let:
                print(f"   ⚪ Row {row_num}: Not qualified ({', '.join(reasons)})")
        
        print(f"\n📊 Processing complete:")
        print(f"   Rows processed: {rows_processed}")
        print(f"   Updates made: {updates_made}")
        
        if updates_made > 0:
            # Save the workbook
            print("💾 Saving changes...")
            workbook.save(file_path)
            print("✅ File saved successfully!")
        else:
            print("ℹ️  No changes needed.")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """
    Main function to run the status updater
    """
    # Default file path
    default_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/30 November 2025 Bank Schedule [updated].xlsx"
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = default_file
        
    # Normalize path
    file_path = os.path.abspath(file_path)
    
    success = update_status_column(file_path)
    
    if success:
        print("\n🎉 Status column update completed successfully!")
    else:
        print("\n❌ Status column update failed.")
        sys.exit(1)

if __name__ == "__main__":
    main()
