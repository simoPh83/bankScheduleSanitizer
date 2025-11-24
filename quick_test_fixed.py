#!/usr/bin/env python3
"""
Quick test for the fixed Legacy View implementation
"""
import os
import openpyxl
import tkinter as tk
from bank_schedule_sanitizer import BankScheduleSanitizer

def test_fixed_implementation():
    # Create root window but don't show it
    root = tk.Tk()
    root.withdraw()  # Hide the window
    
    # Create sanitizer
    sanitizer = BankScheduleSanitizer(root)
    
    # Input and output paths
    input_path = 'data/Leasing Bank Schedule June 2025.xlsx'
    output_path = 'test_fixed_legacy_view.xlsx'
    
    print("🧪 Testing fixed Legacy View implementation...")
    
    try:
        # Analyze the data first
        analysis_results = sanitizer.analyze_bank_schedule_data(input_path)
        print(f"📊 Analysis: {analysis_results['buildings']} buildings, {analysis_results['units']} units")
        
        # Create the output file with the individual methods
        sanitizer.create_buildings_summary_sheet(input_path, output_path, analysis_results)
        sanitizer.create_units_sheet(input_path, output_path, analysis_results['building_names'])
        sanitizer.create_hierarchical_view_sheet_from_files(input_path, output_path, analysis_results['building_names'])
        
        # Check the results
        if os.path.exists(output_path):
            wb = openpyxl.load_workbook(output_path)
            print(f"📋 Created sheets: {wb.sheetnames}")
            
            # Check for Legacy View sheets
            legacy_sheets = [s for s in wb.sheetnames if 'Legacy' in s]
            print(f"🏛️ Legacy View sheets: {legacy_sheets}")
            
            if len(legacy_sheets) == 1 and legacy_sheets[0] == 'Legacy View':
                print("✅ SUCCESS: Single 'Legacy View' sheet created!")
                
                # Check the structure
                ws = wb['Legacy View']
                print(f"✅ Legacy View structure: {ws.max_column} columns × {ws.max_row} rows")
                
                # Check some headers
                headers = []
                for col in range(1, min(11, ws.max_column + 1)):
                    cell_value = ws.cell(1, col).value
                    headers.append(str(cell_value) if cell_value else "")
                print(f"📋 First 10 headers: {headers}")
                
            else:
                print(f"❌ ISSUE: Expected single 'Legacy View' sheet, got: {legacy_sheets}")
            
            wb.close()
            print(f"📁 Output saved to: {output_path}")
            
        else:
            print("❌ ERROR: Output file not created")
            
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
    
    # Clean up
    root.destroy()

if __name__ == "__main__":
    test_fixed_implementation()
