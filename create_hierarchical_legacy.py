#!/usr/bin/env python3

import openpyxl
import os

def get_column_letter(col_num):
    """Convert column number to Excel letter (A, B, C, etc.)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result

def is_numeric_column(header):
    """Check if a column header suggests numeric data that should be summed"""
    if not header:
        return False
    
    header_lower = str(header).lower()
    numeric_keywords = [
        'area', 'rent', 'erv', 'value', 'valuation', 'cap', 'price', 
        'amount', 'total', 'sum', 'cost', '£', '$', 'sq', 'sqft', 
        'sq.ft', 'sq ft', 'square'
    ]
    
    return any(keyword in header_lower for keyword in numeric_keywords)

def create_hierarchical_legacy_view(file_path):
    """Create a hierarchical Legacy View that matches the original Bank Schedule structure"""
    try:
        print(f"Opening file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        
        if "Units" not in workbook.sheetnames:
            print("Error: 'Units' sheet not found")
            return False
            
        units_ws = workbook["Units"]
        
        # Remove existing Legacy View
        if "Legacy View" in workbook.sheetnames:
            del workbook["Legacy View"]
        
        hier_ws = workbook.create_sheet("Legacy View")
        
        # Add title and instructions
        hier_ws.cell(row=1, column=1, value="Legacy View - Hierarchical Building Structure")
        hier_ws.cell(row=2, column=1, value="(Auto-updates from Units sheet - preserves original building grouping)")
        
        # Set up headers in row 4
        headers_row = 4
        
        # Create column mapping and copy headers (skip Building column from Units)
        col_mapping = {}
        target_col = 1
        
        for source_col in range(2, units_ws.max_column + 1):  # Skip column 1 (Building)
            header = units_ws.cell(row=1, column=source_col).value
            if header and str(header).strip():
                header_str = str(header).strip()
                col_mapping[source_col] = target_col
                hier_ws.cell(row=headers_row, column=target_col, value=header_str)
                target_col += 1
        
        print(f"Headers mapped: {len(col_mapping)} columns")
        
        # Group units by building from Units sheet
        buildings = {}
        units_row = units_ws.max_row + 1
        
        for units_data_row in range(2, units_row):
            building_cell = units_ws.cell(row=units_data_row, column=1)
            building_name = building_cell.value
            
            if building_name and str(building_name).strip():
                building_name = str(building_name).strip()
                if building_name not in buildings:
                    buildings[building_name] = []
                buildings[building_name].append(units_data_row)
        
        print(f"Found {len(buildings)} buildings with units:")
        for building, rows in buildings.items():
            print(f"  - {building}: {len(rows)} units")
        
        # Create hierarchical structure
        current_row = headers_row + 1
        total_formulas = 0
        
        # Add empty row after headers
        current_row += 1
        
        for building_name, unit_rows in buildings.items():
            print(f"Processing building: {building_name}")
            
            # Building header row
            building_cell = hier_ws.cell(row=current_row, column=1, value=building_name)
            if hasattr(openpyxl.styles, 'Font'):
                building_cell.font = openpyxl.styles.Font(bold=True, size=12)
            current_row += 1
            
            # Empty row after building header
            current_row += 1
            
            # Unit rows for this building
            building_unit_rows = []
            for units_data_row in unit_rows:
                building_unit_rows.append(current_row)
                
                # Copy unit data with formulas
                for source_col, target_col in col_mapping.items():
                    col_letter = get_column_letter(source_col)
                    formula = f"=Units!{col_letter}{units_data_row}"
                    hier_ws.cell(row=current_row, column=target_col, value=formula)
                    total_formulas += 1
                
                current_row += 1
            
            # Empty row after units
            current_row += 1
            
            # Building summary row with SUM formulas
            summary_cell = hier_ws.cell(row=current_row, column=1, value=f"{building_name} - TOTAL")
            if hasattr(openpyxl.styles, 'Font'):
                summary_cell.font = openpyxl.styles.Font(bold=True, italic=True)
            
            # Add SUM formulas for numeric columns
            for source_col, target_col in col_mapping.items():
                header = units_ws.cell(row=1, column=source_col).value
                if header and is_numeric_column(header):
                    # Create SUM formula for the building's units
                    col_letter = get_column_letter(target_col)
                    start_row = building_unit_rows[0] if building_unit_rows else current_row
                    end_row = building_unit_rows[-1] if building_unit_rows else current_row
                    
                    if start_row <= end_row:
                        sum_formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                        summary_cell_data = hier_ws.cell(row=current_row, column=target_col, value=sum_formula)
                        if hasattr(openpyxl.styles, 'Font'):
                            summary_cell_data.font = openpyxl.styles.Font(bold=True)
                        total_formulas += 1
            
            current_row += 1
            
            # Empty row after building summary
            current_row += 1
        
        # Format headers
        for col in range(1, len(col_mapping) + 1):
            cell = hier_ws.cell(row=headers_row, column=col)
            if hasattr(openpyxl.styles, 'Font'):
                cell.font = openpyxl.styles.Font(bold=True)
        
        # Auto-width columns
        try:
            for column in hier_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                hier_ws.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            print(f"Column auto-width failed: {e}")
        
        # Save the updated file
        workbook.save(file_path)
        workbook.close()
        
        print(f"✅ Hierarchical Legacy View created with {total_formulas} dynamic formulas")
        print("🔗 Building structure preserved with auto-updating formulas and summary totals")
        return True
        
    except Exception as e:
        print(f"❌ Error creating hierarchical Legacy View: {str(e)}")
        return False

if __name__ == "__main__":
    file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/8.xlsx"
    
    if os.path.exists(file_path):
        create_hierarchical_legacy_view(file_path)
    else:
        print(f"File not found: {file_path}")
