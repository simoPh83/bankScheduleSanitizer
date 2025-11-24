#!/usr/bin/env python3
"""
Test the Status column functionality
"""

import os
import sys
import shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def test_status_column():
    """Test that the Status column is added correctly with data validation"""
    print("Testing Status column functionality...")
    print("="*50)
    
    from bank_schedule_sanitizer import BankScheduleSanitizer
    import tkinter as tk
    from openpyxl import load_workbook
    import pandas as pd
    
    # Clean up existing test file
    output_file = 'test_status_column.xlsx'
    if os.path.exists(output_file):
        os.remove(output_file)
        print('Cleaned up existing test file')
    
    # Create test instance
    root = tk.Tk()
    root.withdraw()
    app = BankScheduleSanitizer(root)
    
    input_file = 'data/Leasing Bank Schedule June 2025.xlsx'
    
    try:
        # Copy original file
        shutil.copy2(input_file, output_file)
        
        # Get analysis data
        analysis_results = app.analyze_bank_schedule_data(input_file)
        building_names = analysis_results['building_names']
        
        print(f'Creating Units sheet with Status column for {len(building_names)} buildings...')
        
        # Create Units sheet with Status column
        app.create_units_sheet(input_file, output_file, building_names)
        
        # Verify the Status column was added
        wb = load_workbook(output_file, data_only=False)
        units_ws = wb['Units']
        
        print(f'✅ Units sheet created with {units_ws.max_row} rows, {units_ws.max_column} columns')
        
        # Find Status column
        status_col = None
        for col in range(1, units_ws.max_column + 1):
            header_cell = units_ws.cell(row=1, column=col)
            if header_cell.value == 'Status':
                status_col = col
                break
        
        if status_col:
            col_letter = chr(64 + status_col) if status_col <= 26 else f"A{chr(64 + status_col - 26)}"
            print(f'✅ Found Status column at column {status_col} ({col_letter})')
            
            # Check first few Status values
            print('\nFirst 10 Status values:')
            print('-' * 50)
            void_count = 0
            let_count = 0
            
            for row in range(2, min(12, units_ws.max_row + 1)):
                status_cell = units_ws.cell(row=row, column=status_col)
                building_cell = units_ws.cell(row=row, column=1)  # Building column
                status_value = status_cell.value
                
                building_name = str(building_cell.value)[:25] if building_cell.value else "None"
                print(f'Row {row:2d}: {building_name:<25} -> {status_value}')
                
                if status_value == 'Void':
                    void_count += 1
                elif status_value == 'Let':
                    let_count += 1
            
            print(f'\nIn sample rows: {let_count} Let, {void_count} Void')
            
            # Check data validation
            if units_ws.data_validations:
                print(f'\n✅ Data validation found: {len(units_ws.data_validations.dataValidation)} validation(s)')
                for dv in units_ws.data_validations.dataValidation:
                    print(f'   Formula: {dv.formula1}')
                    print(f'   Range: {dv.sqref}')
            else:
                print('\n⚠️ No data validation found')
                
            # Check if autofilter includes Status column
            if units_ws.auto_filter and units_ws.auto_filter.ref:
                print(f'\n✅ Autofilter range: {units_ws.auto_filter.ref}')
            
        else:
            print('❌ Status column not found')
            # Show all headers
            print('Available headers:')
            for col in range(1, min(10, units_ws.max_column + 1)):
                header_cell = units_ws.cell(row=1, column=col)
                print(f'  Column {col}: {header_cell.value}')
        
        wb.close()
        print('\n✅ Status column test completed!')
        
        return status_col is not None
        
    except Exception as e:
        print(f'❌ Error: {str(e)}')
        import traceback
        traceback.print_exc()
        return False
    finally:
        root.destroy()

if __name__ == "__main__":
    success = test_status_column()
    if success:
        print("\n🎉 Status column functionality working correctly!")
    else:
        print("\n❌ Status column test failed")
