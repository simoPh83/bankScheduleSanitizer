#!/usr/bin/env python3

import openpyxl
import os

def diagnose_excel_file(file_path):
    """Diagnose Excel file structure to help debug VBA issues"""
    try:
        print(f"🔍 Diagnosing Excel file: {file_path}")
        print("=" * 60)
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return
        
        workbook = openpyxl.load_workbook(file_path)
        
        print(f"📋 Sheet Names ({len(workbook.sheetnames)} total):")
        for i, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"  {i}. '{sheet_name}'")
        
        # Check if critical sheets exist
        critical_sheets = ["Units", "Legacy View", "Bank Schedule"]
        print(f"\n🔍 Critical Sheet Check:")
        for sheet in critical_sheets:
            exists = sheet in workbook.sheetnames
            status = "✅" if exists else "❌"
            print(f"  {status} '{sheet}': {'Found' if exists else 'Missing'}")
        
        # Analyze Units sheet if it exists
        if "Units" in workbook.sheetnames:
            print(f"\n📊 Units Sheet Analysis:")
            units_ws = workbook["Units"]
            
            print(f"  📏 Dimensions: {units_ws.max_row} rows × {units_ws.max_column} columns")
            
            # Check headers
            print(f"  📝 Headers (Row 1):")
            for col in range(1, min(units_ws.max_column + 1, 11)):  # Show first 10 columns
                header = units_ws.cell(row=1, column=col).value
                print(f"    Col {col}: '{header}'")
            if units_ws.max_column > 10:
                print(f"    ... and {units_ws.max_column - 10} more columns")
            
            # Check for data
            print(f"  📊 Data Sample (First 5 rows):")
            for row in range(1, min(6, units_ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, units_ws.max_column + 1)):
                    value = units_ws.cell(row=row, column=col).value
                    if value is None:
                        row_data.append("(empty)")
                    else:
                        row_data.append(str(value)[:20])  # Truncate long values
                print(f"    Row {row}: {' | '.join(row_data)}")
            
            # Check building names
            building_count = 0
            buildings = set()
            for row in range(2, min(units_ws.max_row + 1, 102)):  # Check first 100 data rows
                building = units_ws.cell(row=row, column=1).value
                if building and str(building).strip():
                    buildings.add(str(building).strip())
                    building_count += 1
            
            print(f"  🏢 Building Analysis:")
            print(f"    Total rows with building data: {building_count}")
            print(f"    Unique buildings: {len(buildings)}")
            if len(buildings) > 0:
                print(f"    Sample buildings:")
                for i, building in enumerate(list(buildings)[:5]):
                    print(f"      - {building}")
                if len(buildings) > 5:
                    print(f"      ... and {len(buildings) - 5} more")
        
        # Analyze Legacy View sheet if it exists
        if "Legacy View" in workbook.sheetnames:
            print(f"\n📋 Legacy View Sheet Analysis:")
            legacy_ws = workbook["Legacy View"]
            print(f"  📏 Dimensions: {legacy_ws.max_row} rows × {legacy_ws.max_column} columns")
            
            # Check first few rows
            print(f"  📝 Content Sample:")
            for row in range(1, min(6, legacy_ws.max_row + 1)):
                value = legacy_ws.cell(row=row, column=1).value
                if value:
                    print(f"    Row {row}: {str(value)[:50]}")
        
        workbook.close()
        
        print(f"\n✅ Diagnosis complete!")
        print(f"\n💡 VBA Troubleshooting Tips:")
        print(f"  1. Make sure 'Units' sheet exists")
        print(f"  2. Check that Units sheet has data in column A (Building names)")
        print(f"  3. Verify headers are in row 1 of Units sheet")
        print(f"  4. Try running 'TestSheetExists' macro first")
        
    except Exception as e:
        print(f"❌ Error diagnosing file: {str(e)}")

if __name__ == "__main__":
    file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/10_VBA.xlsm"
    diagnose_excel_file(file_path)
