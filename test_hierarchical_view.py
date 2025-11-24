#!/usr/bin/env python3
"""
Test script to verify the Hierarchical View sheet creation functionality
"""

import pandas as pd
import sys
import os
import tempfile
import tkinter as tk

# Add the current directory to path so we can import our main module
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from bank_schedule_sanitizer import BankScheduleSanitizer

def test_hierarchical_view():
    """Test the hierarchical view sheet creation functionality."""
    
    print("Testing Hierarchical View Creation...")
    print("="*60)
    
    test_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/Leasing Bank Schedule June 2025.xlsx"
    
    if not os.path.exists(test_file):
        print(f"❌ Test file not found: {test_file}")
        return False
    
    try:
        # Create a minimal tkinter instance (needed for the class)
        root = tk.Tk()
        root.withdraw()  # Hide the window for testing
        
        # Create an instance of our sanitizer
        sanitizer = BankScheduleSanitizer(root)
        
        # Override the log_message method to print to console for testing
        def test_log(message):
            print(f"[LOG] {message}")
        sanitizer.log_message = test_log
        
        print("📊 Running analysis...")
        
        # Test the analysis function first
        analysis_results = sanitizer.analyze_bank_schedule_data(test_file)
        
        print(f"\n🎉 Analysis completed:")
        print(f"• Buildings: {analysis_results['buildings']}")
        print(f"• Units: {analysis_results['units']}")
        print(f"• Empty rows: {analysis_results['empty_rows']}")
        print(f"• Building names: {len(analysis_results['building_names'])} found")
        
        # Create a temporary output file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            output_path = temp_file.name
        
        print(f"\n📋 Creating sanitized file: {os.path.basename(output_path)}")
        
        # Create Buildings sheet
        print("📊 Creating Buildings sheet...")
        sanitizer.create_buildings_summary_sheet(test_file, output_path, analysis_results)
        
        # Create Units sheet  
        print("📋 Creating Units sheet...")
        sanitizer.create_units_sheet(test_file, output_path, analysis_results['building_names'])
        
        # Create Hierarchical View sheet
        print("🏢 Creating Hierarchical View sheet...")
        success = sanitizer.create_hierarchical_view_sheet_from_files(test_file, output_path, analysis_results['building_names'])
        
        if success:
            print(f"\n✅ All sheets created successfully!")
            print(f"📁 Output file: {output_path}")
            
            # Verify the sheets exist
            import openpyxl
            workbook = openpyxl.load_workbook(output_path)
            sheets = workbook.sheetnames
            print(f"📊 Sheets created: {sheets}")
            
            if 'Hierarchical View' in sheets:
                hier_ws = workbook['Hierarchical View']
                print(f"🏢 Hierarchical View sheet has {hier_ws.max_row} rows and {hier_ws.max_column} columns")
                
                # Show first few rows
                print(f"\n📋 Sample content from Hierarchical View:")
                for row in range(1, min(10, hier_ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(5, hier_ws.max_column + 1)):
                        cell_value = hier_ws.cell(row=row, column=col).value
                        if cell_value:
                            row_data.append(str(cell_value)[:30])
                    if row_data:
                        print(f"   Row {row}: {' | '.join(row_data)}")
            
            workbook.close()
            
            # Clean up
            os.unlink(output_path)
            print(f"\n🧹 Cleaned up temporary file")
            return True
            
        else:
            print(f"\n❌ Failed to create Hierarchical View sheet")
            # Clean up
            os.unlink(output_path)
            return False
        
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_hierarchical_view()
    if success:
        print(f"\n🎉 Hierarchical View test completed successfully!")
    else:
        print(f"\n💥 Hierarchical View test failed!")
        sys.exit(1)
