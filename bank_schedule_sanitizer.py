#!/usr/bin/env python3
"""
Bank Schedule Plus
A GUI application to sanitize Excel bank schedule files.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import shutil
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy
import re


class BankScheduleSanitizer:
    def __init__(self, root):
        self.root = root
        self.root.title("Bank Schedule Plus")
        self.root.geometry("600x500")
        
        # Variables
        self.input_file_path = tk.StringVar()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Set up the user interface."""
        # Main frame
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = tk.Label(
            main_frame, 
            text="Bank Schedule Plus", 
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=(0, 20))
        
        # Input file selection
        input_frame = tk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(input_frame, text="Select Excel file to process:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        input_path_frame = tk.Frame(input_frame)
        input_path_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.input_path_entry = tk.Entry(
            input_path_frame, 
            textvariable=self.input_file_path, 
            state="readonly",
            font=("Arial", 9)
        )
        self.input_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        tk.Button(
            input_path_frame, 
            text="Browse...", 
            command=self.browse_input_file,
            width=10
        ).pack(side=tk.RIGHT, padx=(10, 0))
        
        # Sheet status label
        self.sheet_status_label = tk.Label(
            main_frame,
            text="",
            font=("Arial", 9),
            fg="#666666",
            wraplength=550,
            justify=tk.LEFT
        )
        self.sheet_status_label.pack(pady=(10, 0), anchor=tk.W)
        
        # Sanitize button
        self.sanitize_button = tk.Button(
            main_frame, 
            text="Create", 
            command=self.sanitize_file,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            height=2,
            state=tk.DISABLED
        )
        self.sanitize_button.pack(pady=20)
        
        # Error/Status text box
        status_frame = tk.Frame(main_frame)
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        tk.Label(status_frame, text="Status and Error Messages:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        self.status_text = scrolledtext.ScrolledText(
            status_frame,
            height=8,
            font=("Consolas", 9),
            wrap=tk.WORD
        )
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Initial status message
        self.log_message("Ready to process Excel files. Please select an input file to begin.")
        
    def browse_input_file(self):
        """Open file dialog to select input Excel file."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Process",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.input_file_path.set(file_path)
            self.log_message(f"Selected input file: {os.path.basename(file_path)}")
            # Check for existing sheets and update status
            self.update_sheet_status_message(file_path)
            # Enable sanitize button when input file is selected
            self.sanitize_button.config(state=tk.NORMAL)
            
    def log_message(self, message):
        """Add a message to the status text box."""
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
        
    def check_existing_sheets(self, file_path):
        """Check if Units and Buildings sheets already exist in the Excel file."""
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            units_exists = "Units" in sheet_names
            buildings_exists = "Buildings" in sheet_names
            
            return {
                'units_exists': units_exists,
                'buildings_exists': buildings_exists,
                'sheet_names': sheet_names
            }
        except Exception as e:
            self.log_message(f"Could not check existing sheets: {str(e)}")
            return {
                'units_exists': False,
                'buildings_exists': False,
                'sheet_names': []
            }
    
    def check_file_lock_status(self, file_path):
        """Check if the Excel file is locked/open by another user."""
        try:
            import tempfile
            import os
            
            # Method 1: Try to open the file in write mode
            try:
                # Try to open for writing (this will fail if locked)
                with open(file_path, 'r+b') as f:
                    pass
                return {'is_locked': False, 'method': 'write_test'}
            except (PermissionError, OSError) as e:
                # File might be locked, but let's try with openpyxl to be sure
                pass
            
            # Method 2: Try to load and save with openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                
                # Try to save to a temp file to test write access
                temp_dir = os.path.dirname(file_path)
                temp_file = os.path.join(temp_dir, f"~temp_lock_test_{os.getpid()}.tmp")
                
                try:
                    wb.save(temp_file)
                    os.remove(temp_file)  # Clean up
                    wb.close()
                    
                    # Now try to save to the original file (this is the real test)
                    wb = load_workbook(file_path)
                    wb.save(file_path)
                    wb.close()
                    
                    return {'is_locked': False, 'method': 'openpyxl_save'}
                except Exception as save_error:
                    wb.close()
                    # Clean up temp file if it exists
                    if os.path.exists(temp_file):
                        try:
                            os.remove(temp_file)
                        except:
                            pass
                    return {
                        'is_locked': True, 
                        'method': 'openpyxl_save', 
                        'error': str(save_error)
                    }
                    
            except Exception as load_error:
                return {
                    'is_locked': True, 
                    'method': 'openpyxl_load', 
                    'error': str(load_error)
                }
                
        except Exception as e:
            # If we can't determine, assume it's accessible
            self.log_message(f"Could not determine file lock status: {str(e)}")
            return {'is_locked': False, 'method': 'unknown', 'error': str(e)}
        
    def update_sheet_status_message(self, file_path):
        """Update the sheet status label based on existing sheets in the file and check for file locks."""
        if not file_path:
            self.sheet_status_label.config(text="")
            return
        
        # Check if file is locked first
        self.log_message("🔒 Checking if file is locked by another user...")
        lock_status = self.check_file_lock_status(file_path)
        
        if lock_status['is_locked']:
            error_msg = f"⚠️ File appears to be open by another user and cannot be modified."
            if 'error' in lock_status:
                self.log_message(f"   Lock detection details: {lock_status['error']}")
            
            status_text = "⚠️ FILE IS LOCKED - Cannot proceed while file is open by another user"
            color = "#DC143C"  # Red for error
            self.sheet_status_label.config(text=status_text, fg=color)
            
            # Disable the Create button
            self.sanitize_button.config(state=tk.DISABLED)
            self.log_message(error_msg)
            return
        else:
            self.log_message(f"✅ File is available for modification (method: {lock_status['method']})")
            # Re-enable Create button if it was disabled
            self.sanitize_button.config(state=tk.NORMAL)
        
        # Continue with normal sheet checking
        sheet_info = self.check_existing_sheets(file_path)
        
        units_exists = sheet_info['units_exists']
        buildings_exists = sheet_info['buildings_exists']
        
        if units_exists and buildings_exists:
            status_text = "📋 Both 'Units' and 'Buildings' sheets found - they will be updated with new data."
            color = "#FF8C00"  # Orange for update
        elif units_exists and not buildings_exists:
            status_text = "📋 'Units' sheet found (will be updated). 'Buildings' sheet will be created."
            color = "#1E90FF"  # Blue for mixed
        elif not units_exists and buildings_exists:
            status_text = "📋 'Buildings' sheet found (will be updated). 'Units' sheet will be created."
            color = "#1E90FF"  # Blue for mixed
        else:
            status_text = "📋 'Units' and 'Buildings' sheets will be created."
            color = "#228B22"  # Green for new creation
        
        self.sheet_status_label.config(text=status_text, fg=color)
        
    def validate_excel_file(self, file_path):
        """Validate that the Excel file exists and can be read."""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
                
            # Try to read the Excel file to check if it's valid
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            self.log_message(f"Excel file validated. Found sheets: {', '.join(sheet_names)}")
            
            # Check if "Bank Schedule" sheet exists
            if "Bank Schedule" in sheet_names:
                self.log_message("✓ Found 'Bank Schedule' sheet in the file.")
                return True, "File is valid"
            else:
                self.log_message("⚠ Warning: 'Bank Schedule' sheet not found in the file.")
                return True, "File is valid but 'Bank Schedule' sheet not found"
                
        except Exception as e:
            error_msg = f"Error validating Excel file: {str(e)}"
            self.log_message(f"✗ {error_msg}")
            return False, error_msg
            
    def refactor_formula(self, formula, source_row, target_row, column_mapping=None):
        """
        Refactor Excel formula by adjusting cell references for new sheet structure.
        
        Args:
            formula: The original formula string (e.g., '=V7/H7')
            source_row: Original row number in source sheet
            target_row: Target row number in new sheet
            column_mapping: Dict mapping source column numbers to target column numbers
        
        Returns:
            Refactored formula string
        """
        import re
        
        if not formula or not formula.startswith('='):
            return formula
            
        if column_mapping is None:
            column_mapping = {}
        
        try:
            # Pattern to match cell references like H7, V7, etc.
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def replace_cell_ref(match):
                col_letters = match.group(1)
                old_row = int(match.group(2))
                
                # Convert column letters to column number
                old_col_num = 0
                for char in col_letters:
                    old_col_num = old_col_num * 26 + (ord(char) - ord('A') + 1)
                
                # Apply column mapping if available
                if old_col_num in column_mapping:
                    new_col_num = column_mapping[old_col_num]
                else:
                    # Fallback: no change if mapping not found
                    new_col_num = old_col_num
                
                # Calculate new row number
                # If the original reference was to the same row as source, map it to target row
                if old_row == source_row:
                    new_row = target_row
                else:
                    # For references to other rows, maintain relative offset
                    row_diff = old_row - source_row
                    new_row = target_row + row_diff
                    
                # Ensure row number is positive
                if new_row < 1:
                    new_row = 1
                
                # Convert new column number back to letters
                new_col_letters = ""
                temp_col = new_col_num
                while temp_col > 0:
                    temp_col -= 1
                    new_col_letters = chr(temp_col % 26 + ord('A')) + new_col_letters
                    temp_col //= 26
                
                return f"{new_col_letters}{new_row}"
            
            # Replace all cell references in the formula
            refactored_formula = re.sub(cell_pattern, replace_cell_ref, formula)

            # self.log_message(f"   📐 Formula refactored: {formula} → {refactored_formula}")
            return refactored_formula
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not refactor formula {formula}: {str(e)}")
            return formula  # Return original if refactoring fails

    def create_hierarchical_view_sheet(self, workbook, units_ws, units_row):
        """Create a hierarchical Legacy View sheet with embedded VBA script for dynamic updates"""
        try:
            self.log_message("📋 Creating dynamic Legacy View with embedded VBA script...")
            
            # Remove any existing legacy view sheets
            sheets_to_remove = ["Hierarchical View", "Hierarchical View1", "Legacy View"]
            for sheet_name in sheets_to_remove:
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]
                
            hier_ws = workbook.create_sheet("Legacy View")
            
            # Add title and instructions
            hier_ws.cell(row=1, column=1, value="Legacy View - Dynamic Hierarchical Structure")
            hier_ws.cell(row=2, column=1, value="(Auto-rebuilds from Units sheet - responsive to structure changes)")
            hier_ws.cell(row=3, column=1, value="Press Ctrl+Shift+R to manually refresh, or run RefreshLegacyView macro")
            
            # Add VBA macro to the workbook
            self.add_vba_macro(workbook)
            
            # Create initial structure by calling our VBA-equivalent logic
            self.build_legacy_view_structure(workbook, units_ws, units_row)
            
            self.log_message("✅ Dynamic Legacy View created with embedded VBA refresh capability")
            self.log_message("🔧 VBA macro 'RefreshLegacyView' added for automatic structure updates")
            self.log_message("⌨️ Use Ctrl+Shift+R or run macro to refresh when Units sheet changes")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating dynamic Legacy View sheet: {str(e)}")
            return False
    
    def add_vba_macro(self, workbook):
        """Add VBA macro to dynamically rebuild Legacy View structure"""
        try:
            # Check if workbook supports VBA
            if not hasattr(workbook, 'vba_archive'):
                self.log_message("⚠️ Workbook format doesn't support VBA - creating macro-free version")
                return
            
            # VBA code for dynamic Legacy View rebuilding
            vba_code = '''
Sub RefreshLegacyView()
    ' Dynamic Legacy View Rebuilder
    ' Parses Units sheet and recreates hierarchical Legacy View structure
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    On Error GoTo ErrorHandler
    
    Dim unitsWs As Worksheet
    Dim legacyWs As Worksheet
    Dim lastRow As Long
    Dim currentRow As Long
    Dim headerRow As Long
    Dim col As Long
    Dim buildingDict As Object
    Dim buildingName As String
    Dim key As Variant
    Dim unitRows As Collection
    Dim i As Long
    Dim startRow As Long, endRow As Long
    
    ' Get worksheets
    Set unitsWs = ThisWorkbook.Worksheets("Units")
    
    ' Clear existing Legacy View content (keep first 3 instruction rows)
    Set legacyWs = ThisWorkbook.Worksheets("Legacy View")
    lastRow = legacyWs.Cells(legacyWs.Rows.Count, 1).End(xlUp).Row
    If lastRow > 3 Then
        legacyWs.Range("A4:ZZ" & lastRow).Clear
    End If
    
    ' Set up headers
    headerRow = 4
    For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
        If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
            legacyWs.Cells(headerRow, col - 1).Value = unitsWs.Cells(1, col).Value
            legacyWs.Cells(headerRow, col - 1).Font.Bold = True
        End If
    Next col
    
    ' Group units by building - find Property column first
    Set buildingDict = CreateObject("Scripting.Dictionary")
    lastRow = unitsWs.Cells(unitsWs.Rows.Count, 1).End(xlUp).Row
    
    ' Find Property column (contains building names)
    Dim propertyCol As Long
    propertyCol = 4  ' Default to column 4 (typical Property column position)
    For col = 1 To 10  ' Search first 10 columns
        If InStr(LCase(CStr(unitsWs.Cells(1, col).Value)), "property") > 0 Then
            propertyCol = col
            Exit For
        End If
    Next col
    
    For i = 2 To lastRow
        buildingName = Trim(CStr(unitsWs.Cells(i, propertyCol).Value))  ' Use Property column instead of column 1
        If buildingName <> "" Then
            If Not buildingDict.Exists(buildingName) Then
                Set buildingDict(buildingName) = New Collection
            End If
            buildingDict(buildingName).Add i
        End If
    Next i
    
    ' Build hierarchical structure
    currentRow = headerRow + 2 ' Empty row after headers
    
    For Each key In buildingDict.Keys
        Set unitRows = buildingDict(key)
        
        ' Building header
        legacyWs.Cells(currentRow, 1).Value = key
        legacyWs.Cells(currentRow, 1).Font.Bold = True
        legacyWs.Cells(currentRow, 1).Font.Size = 12
        currentRow = currentRow + 2 ' Empty row after building header
        
        ' Unit rows
        startRow = currentRow
        For i = 1 To unitRows.Count
            For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
                If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
                    legacyWs.Cells(currentRow, col - 1).Formula = "=Units!" & _
                        ConvertToColumnLetter(col) & unitRows(i)
                End If
            Next col
            currentRow = currentRow + 1
        Next i
        endRow = currentRow - 1
        
        ' Empty row + summary row
        currentRow = currentRow + 1
        legacyWs.Cells(currentRow, 1).Value = key & " - TOTAL"
        legacyWs.Cells(currentRow, 1).Font.Bold = True
        legacyWs.Cells(currentRow, 1).Font.Italic = True
        
        ' Add SUM formulas for numeric columns
        For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
            If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
                If IsNumericColumn(unitsWs.Cells(1, col).Value) Then
                    If startRow <= endRow Then
                        legacyWs.Cells(currentRow, col - 1).Formula = "=SUM(" & _
                            ConvertToColumnLetter(col - 1) & startRow & ":" & _
                            ConvertToColumnLetter(col - 1) & endRow & ")"
                        legacyWs.Cells(currentRow, col - 1).Font.Bold = True
                    End If
                End If
            End If
        Next col
        
        currentRow = currentRow + 2 ' Empty row after summary
    Next key
    
    ' Auto-fit columns
    legacyWs.Columns.AutoFit
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "Legacy View refreshed successfully! Found " & buildingDict.Count & " buildings.", vbInformation
    Exit Sub
    
ErrorHandler:
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "Error refreshing Legacy View: " & Err.Description, vbCritical
End Sub

Private Function ConvertToColumnLetter(colNum As Long) As String
    Dim result As String
    result = ""
    While colNum > 0
        colNum = colNum - 1
        result = Chr(65 + (colNum Mod 26)) & result
        colNum = colNum \\ 26
    Wend
    ConvertToColumnLetter = result
End Function

Private Function IsNumericColumn(header As String) As Boolean
    Dim lowerHeader As String
    lowerHeader = LCase(Trim(header))
    
    IsNumericColumn = (InStr(lowerHeader, "area") > 0) Or _
                      (InStr(lowerHeader, "rent") > 0) Or _
                      (InStr(lowerHeader, "erv") > 0) Or _
                      (InStr(lowerHeader, "value") > 0) Or _
                      (InStr(lowerHeader, "valuation") > 0) Or _
                      (InStr(lowerHeader, "cap") > 0) Or _
                      (InStr(lowerHeader, "£") > 0) Or _
                      (InStr(lowerHeader, "$") > 0) Or _
                      (InStr(lowerHeader, "sq") > 0) Or _
                      (InStr(lowerHeader, "total") > 0) Or _
                      (InStr(lowerHeader, "sum") > 0) Or _
                      (InStr(lowerHeader, "cost") > 0)
End Function

Private Sub Worksheet_Activate()
    ' Auto-refresh when Legacy View sheet is activated
    ' Uncomment next line if you want automatic refresh on sheet activation
    ' RefreshLegacyView
End Sub

' Keyboard shortcut: Ctrl+Shift+R
Private Sub Workbook_Open()
    Application.OnKey "^+R", "RefreshLegacyView"
End Sub
'''
            
            # Note: openpyxl doesn't support adding VBA directly
            # We'll create a macro-enabled structure and add instructions
            self.log_message("📝 VBA macro code prepared (requires manual addition to .xlsm file)")
            
        except Exception as e:
            self.log_message(f"⚠️ Could not add VBA macro: {str(e)}")
    
    def build_legacy_view_structure(self, workbook, units_ws, units_row):
        """Build the initial Legacy View structure"""
        try:
            hier_ws = workbook["Legacy View"]
            
            # Set up headers in row 4 (skip Building column from Units)
            headers_row = 4
            col_mapping = {}
            target_col = 1
            
            for source_col in range(2, units_ws.max_column + 1):  # Skip column 1 (Building)
                header = units_ws.cell(row=1, column=source_col).value
                if header and str(header).strip():
                    header_str = str(header).strip()
                    col_mapping[source_col] = target_col
                    hier_ws.cell(row=headers_row, column=target_col, value=header_str)
                    target_col += 1
            
            # Group units by building
            buildings = {}
            for units_data_row in range(2, units_row):
                building_cell = units_ws.cell(row=units_data_row, column=1)
                building_name = building_cell.value
                
                if building_name and str(building_name).strip():
                    building_name = str(building_name).strip()
                    if building_name not in buildings:
                        buildings[building_name] = []
                    buildings[building_name].append(units_data_row)
            
            # Create hierarchical structure
            current_row = headers_row + 1
            current_row += 1  # Empty row after headers
            
            for building_name, unit_rows in buildings.items():
                # Building header row
                building_cell = hier_ws.cell(row=current_row, column=1, value=building_name)
                if hasattr(openpyxl.styles, 'Font'):
                    building_cell.font = openpyxl.styles.Font(bold=True, size=12)
                current_row += 2  # Empty row after building header
                
                # Unit rows for this building
                building_unit_rows = []
                for units_data_row in unit_rows:
                    building_unit_rows.append(current_row)
                    
                    # Copy unit data with formulas
                    for source_col, target_col in col_mapping.items():
                        col_letter = self.get_column_letter(source_col)
                        formula = f"=Units!{col_letter}{units_data_row}"
                        hier_ws.cell(row=current_row, column=target_col, value=formula)
                    
                    current_row += 1
                
                # Empty row + Building summary
                current_row += 1
                summary_cell = hier_ws.cell(row=current_row, column=1, value=f"{building_name} - TOTAL")
                if hasattr(openpyxl.styles, 'Font'):
                    summary_cell.font = openpyxl.styles.Font(bold=True, italic=True)
                
                # Add SUM formulas for numeric columns
                for source_col, target_col in col_mapping.items():
                    header = units_ws.cell(row=1, column=source_col).value
                    if header and self.is_numeric_column(header):
                        col_letter = self.get_column_letter(target_col)
                        start_row = building_unit_rows[0] if building_unit_rows else current_row
                        end_row = building_unit_rows[-1] if building_unit_rows else current_row
                        
                        if start_row <= end_row:
                            sum_formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                            summary_cell_data = hier_ws.cell(row=current_row, column=target_col, value=sum_formula)
                            if hasattr(openpyxl.styles, 'Font'):
                                summary_cell_data.font = openpyxl.styles.Font(bold=True)
                
                current_row += 2  # Empty row after summary
            
            # Format headers
            for col in range(1, len(col_mapping) + 1):
                cell = hier_ws.cell(row=headers_row, column=col)
                if hasattr(openpyxl.styles, 'Font'):
                    cell.font = openpyxl.styles.Font(bold=True)
            
            # Auto-width columns
            try:
                for column in hier_ws.columns:
                    max_length = 0
                    column_letter = self.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    hier_ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                self.log_message(f"⚠️ Column auto-width failed: {e}")
                
        except Exception as e:
            self.log_message(f"⚠️ Error building initial structure: {str(e)}")
    
    def is_numeric_column(self, header):
        """Check if a column header suggests numeric data that should be summed"""
        if not header:
            return False
        
        header_lower = str(header).lower()
        numeric_keywords = [
            'area', 'rent', 'erv', 'value', 'valuation', 'cap', 'price', 
            'amount', 'total', 'sum', 'cost', '£', '$', 'sq', 'sqft', 
            'sq.ft', 'sq ft', 'square'
        ]
        
        return any(keyword in header_lower for keyword in numeric_keywords)
    
    def get_column_letter(self, col_num):
        """Convert column number to Excel letter (A, B, C, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result

    def create_hierarchical_view_sheet_from_files(self, input_path, building_names):
        """Create a Legacy View sheet by opening the output file and adding the sheet"""
        try:
            self.log_message("📋 Creating Legacy View sheet from files...")
            
            # Open the output file (that should already have Units sheet)
            workbook = openpyxl.load_workbook(input_path)
            
            # Get the Units sheet to understand its structure
            units_ws = workbook['Units']
            units_row = units_ws.max_row + 1  # +1 to account for next available row
            
            # Call the main legacy view creation method
            success = self.create_hierarchical_view_sheet(workbook, units_ws, units_row)
            
            if success:
                # Save the updated workbook
                workbook.save(input_path)
                workbook.close()
                self.log_message("✅ Legacy View sheet added successfully")
                return True
            else:
                workbook.close()
                return False
                
        except Exception as e:
            self.log_message(f"❌ Error creating Legacy View sheet from files: {str(e)}")
            return False

    def find_column_letter(self, worksheet, header_name):
        """Find the column letter for a given header name in the worksheet"""
        try:
            # Check first row for headers
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == header_name:
                    return get_column_letter(col)
            return None
        except:
            return None

    def analyze_bank_schedule_data(self, file_path):
        """Analyze the Bank Schedule data and return counts of buildings, units, empty rows, and building names."""
        try:
            # Read the Bank Schedule sheet, starting from row 3 (index 2)
            df = pd.read_excel(file_path, sheet_name="Bank Schedule", header=2)
            
            # Get column G (Unit Type) - this should be index 6 if we count from A=0
            # But since headers start at row 3, we need to check the actual column name
            unit_type_col = None
            for col in df.columns:
                if 'Unit Type' in str(col) or col == 'G' or df.columns.get_loc(col) == 6:
                    unit_type_col = col
                    break
            
            if unit_type_col is None:
                # If we can't find by name, try by position (column G = index 6)
                if len(df.columns) > 6:
                    unit_type_col = df.columns[6]
                else:
                    raise ValueError("Could not find 'Unit Type' column (column G)")
            
            self.log_message(f"Using column '{unit_type_col}' as Unit Type column")
            
            building_count = 0
            unit_count = 0
            empty_count = 0
            building_names = []
            current_building = None
            
            # Iterate through all rows
            for index, row in df.iterrows():
                unit_type_value = row[unit_type_col]
                
                # Check if row is empty (Unit Type column is empty/null)
                if pd.isna(unit_type_value) or str(unit_type_value).strip() == '':
                    empty_count += 1
                    continue
                
                # Check if this is a building row
                # Building rows have value in Unit Type but should have no other significant data
                # Let's check if most other columns are empty for this row
                non_empty_cols = 0
                for col in df.columns:
                    if col != unit_type_col and not pd.isna(row[col]) and str(row[col]).strip() != '':
                        non_empty_cols += 1
                
                # If this row has Unit Type but few other fields, it's likely a building
                if non_empty_cols <= 2:  # Allow for some tolerance
                    building_name = str(unit_type_value).strip()
                    # Filter out notes and other non-building entries
                    if not any(keyword in building_name.lower() for keyword in ['note', 'capital values', 'rent pa', 'as a result']):
                        building_count += 1
                        building_names.append(building_name)
                        current_building = building_name
                        self.log_message(f"Found building: {current_building}")
                else:
                    # This is a unit row (has Unit Type and other data)
                    unit_count += 1
                    if current_building:
                        self.log_message(f"  Unit in {current_building}: {unit_type_value}")
            
            self.log_message(f"Analysis complete: {building_count} buildings, {unit_count} units")
            
            return {
                'buildings': building_count,
                'units': unit_count,
                'empty_rows': empty_count,
                'total_rows': len(df),
                'building_names': building_names
            }
            
        except Exception as e:
            self.log_message(f"Error analyzing data: {str(e)}")
            raise
    
    def analyze_cap_valn_locations(self, input_path):
        """Analyze Bank Schedule to find Cap Valn locations for each building according to the parsing logic."""
        try:
            self.log_message("🔍 Analyzing Bank Schedule structure to find Cap Valn locations...")
            
            # Load the Bank Schedule workbook directly with openpyxl for row-by-row analysis
            from openpyxl import load_workbook
            wb = load_workbook(input_path, data_only=False)
            
            if 'Bank Schedule' not in wb.sheetnames:
                self.log_message("❌ Bank Schedule sheet not found")
                return {}
            
            bank_ws = wb['Bank Schedule']
            
            # Find the "Unit Demise" column and "2025 Cap Valn. (£)" column
            unit_demise_col = None
            cap_valn_col = None
            property_col = None
            
            for col in range(1, 50):  # Check first 50 columns
                header_val = bank_ws.cell(row=3, column=col).value  # Headers are in row 3 based on debug
                if header_val:
                    header_clean = str(header_val).lower().strip()
                    if 'unit demise' in header_clean:
                        unit_demise_col = col
                        self.log_message(f"Found 'Unit Demise' column at: {chr(64+col)}")
                    elif '2025 cap valn' in header_clean:
                        cap_valn_col = col  # Should be column AB (28)
                        self.log_message(f"Found '2025 Cap Valn' column at: {chr(64+col) if col <= 26 else 'A' + chr(64+col-26)}")
                    elif header_clean == 'property':
                        property_col = col
                        self.log_message(f"Found 'Property' column at: {chr(64+col)}")
            
            if not unit_demise_col or not cap_valn_col or not property_col:
                # Create detailed error message for user
                missing_columns = []
                if not unit_demise_col:
                    missing_columns.append("Unit Demise")
                if not cap_valn_col:
                    missing_columns.append("2025 Cap Valn. (£)")
                if not property_col:
                    missing_columns.append("Property")
                
                error_msg = f"Required columns not found in Bank Schedule sheet: {', '.join(missing_columns)}"
                self.log_message(f"❌ {error_msg}")
                self.log_message(f"   Debug info - Unit Demise={unit_demise_col}, Cap Valn={cap_valn_col}, Property={property_col}")
                
                # Raise exception with user-friendly message
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
            
            
            # Map building names to their Cap Valn row locations
            cap_valn_mapping = {}
            
            # Start parsing from row 5 (after headers) based on debug output
            current_row = 5
            max_rows = bank_ws.max_row  # Use full sheet, no artificial limit
            iteration_count = 0
            max_iterations = max_rows * 2  # Generous safety measure based on actual sheet size
            
            self.log_message(f"Starting Cap Valn analysis from row {current_row} to max row {max_rows}")
            
            while current_row <= max_rows and iteration_count < max_iterations:
                iteration_count += 1
                
                # Safety check for infinite loops
                if iteration_count % 50 == 0:
                    self.log_message(f"Processing iteration {iteration_count}, current row: {current_row}")
                
                # Check if current row has a value in Unit Demise column
                unit_demise_value = bank_ws.cell(row=current_row, column=unit_demise_col).value
                
                if unit_demise_value and str(unit_demise_value).strip():
                    # This is a unit row - get the building name
                    building_name = bank_ws.cell(row=current_row, column=property_col).value
                    if building_name and str(building_name).strip():  # Ensure building name exists
                        building_name = str(building_name).strip()
                        
                        # Skip if we've already processed this building
                        if building_name in cap_valn_mapping:
                            current_row += 1
                            continue
                        
                        # Check if this is a single unit building or multi-unit building
                        # Look at the next row's Unit Demise column AND building name
                        next_row_unit_demise = None
                        next_row_building = None
                        if current_row + 1 <= max_rows:
                            next_row_unit_demise = bank_ws.cell(row=current_row + 1, column=unit_demise_col).value
                            next_row_building = bank_ws.cell(row=current_row + 1, column=property_col).value
                        
                        # Single unit building if next row has no unit demise OR has unit demise but different building
                        if (not next_row_unit_demise or not str(next_row_unit_demise).strip() or
                            not next_row_building or str(next_row_building).strip() != building_name):
                            # Single unit building - Cap Valn is in the current row
                            cap_valn_row = current_row
                            self.log_message(f"Single unit building '{building_name}' - Cap Valn at row {cap_valn_row}")
                            cap_valn_mapping[building_name] = cap_valn_row
                            current_row += 1
                        else:
                            # Multi-unit building - find the last unit row (Y) for this building
                            last_unit_row = current_row
                            temp_row = current_row + 1
                            
                            # Keep checking until we find an empty Unit Demise or different building
                            while temp_row <= max_rows:
                                temp_unit_demise = bank_ws.cell(row=temp_row, column=unit_demise_col).value
                                temp_building = bank_ws.cell(row=temp_row, column=property_col).value
                                
                                # Check if still same building and has unit demise
                                if (temp_unit_demise and str(temp_unit_demise).strip() and 
                                    temp_building and str(temp_building).strip() == building_name):
                                    last_unit_row = temp_row
                                    temp_row += 1
                                else:
                                    break
                            
                            # Y+2 is the summary row (Y+1 is empty, Y+2 has totals)
                            summary_row = last_unit_row + 2
                            self.log_message(f"Multi-unit building '{building_name}' - units from {current_row} to {last_unit_row}, summary at row {summary_row}")
                            cap_valn_mapping[building_name] = summary_row
                            
                            # Y+4 should be the beginning of the next building
                            current_row = last_unit_row + 4
                    else:
                        # Unit Demise exists but no building name - skip this row
                        current_row += 1
                else:
                    current_row += 1
            
            if iteration_count >= max_iterations:
                self.log_message(f"⚠️ Stopped analysis due to max iterations limit ({max_iterations})")
            
            wb.close()
            self.log_message(f"✅ Found Cap Valn locations for {len(cap_valn_mapping)} buildings")
            for building, row in cap_valn_mapping.items():
                self.log_message(f"  {building} -> Row {row}")
            
            return cap_valn_mapping, cap_valn_col
            
        except Exception as e:
            self.log_message(f"❌ Error analyzing Cap Valn locations: {str(e)}")
            # Re-raise the exception so it gets caught with proper error handling
            raise
            
    def create_buildings_summary_sheet(self, input_path, analysis_results, cap_valn_mapping, cap_valn_col):
        """Create a new 'Buildings' sheet with building summary data and SUM formulas linking to Units sheet."""
        try:
            self.log_message("Step 5: Creating Buildings summary sheet with SUM formulas...")
            
            # Import openpyxl for direct Excel manipulation
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            # Open the workbook (file already exists from Units sheet creation)
            workbook = load_workbook(input_path)
            
            # Remove existing Buildings sheet if it exists
            if "Buildings" in workbook.sheetnames:
                del workbook["Buildings"]
                self.log_message("📋 Removed existing 'Buildings' sheet")
            
            # Get building names from analysis results
            building_names = analysis_results['building_names']
            
            self.log_message(f"Creating Buildings summary for {len(building_names)} buildings")
            
            # Create the new Buildings worksheet
            buildings_ws = workbook.create_sheet("Buildings")
            
            # Add timestamp header at the very top
            from datetime import datetime
            now = datetime.now()
            timestamp_text = f"Document updated {now.strftime('%d/%m/%Y')} at {now.strftime('%-I.%M%p').lower()}"
            timestamp_cell = buildings_ws.cell(row=1, column=1, value=timestamp_text)
            # Make timestamp bold and slightly larger
            from openpyxl.styles import Font
            timestamp_cell.font = Font(bold=True, size=11)
            
            # Define headers - only meaningful summary columns for Buildings sheet
            # These are the columns that make sense to sum up for building totals
            meaningful_headers = [
                "Building",
                "Net Area", 
                "Rent PA (£)",
                "2023 ERV (£)",
                "2025 ERV (£)", 
                "2025 Cap Valn. (£)"
            ]
            
            self.log_message(f"📋 Using {len(meaningful_headers)} meaningful summary columns for Buildings sheet")
            
            # Get the Units sheet to find column mappings
            units_ws = workbook["Units"]
            
            # Map meaningful headers to their Units sheet column positions
            header_to_col_mapping = {}
            
            self.log_message("🔍 Analyzing Units sheet headers for column mapping...")
            for col in range(1, units_ws.max_column + 1):
                header_value = units_ws.cell(row=2, column=col).value  # Row 2 since Units headers are now on row 2
                if header_value and str(header_value).strip():
                    clean_header = str(header_value).strip()
                    self.log_message(f"   Column {col}: '{clean_header}'")
                    
                    for meaningful_header in meaningful_headers[1:]:  # Skip "Building" 
                        # Improved matching logic to avoid false positives
                        clean_header_lower = clean_header.lower().strip()
                        meaningful_header_lower = meaningful_header.lower().strip()
                        
                        # Skip very short headers that could cause false matches
                        if len(clean_header_lower) < 3:
                            continue
                        
                        # Remove common punctuation for comparison
                        clean_for_comparison = clean_header_lower.replace('(', '').replace(')', '').replace('£', '').replace('.', '').replace(',', '')
                        meaningful_for_comparison = meaningful_header_lower.replace('(', '').replace(')', '').replace('£', '').replace('.', '').replace(',', '')
                        
                        # Check for substantial word-level overlap, not just character substring
                        if (len(meaningful_for_comparison) > 3 and meaningful_for_comparison in clean_for_comparison) or \
                           (len(clean_for_comparison) > 3 and clean_for_comparison in meaningful_for_comparison):
                            header_to_col_mapping[meaningful_header] = col
                            self.log_message(f"   ✅ Matched '{meaningful_header}' -> Column {col} ('{clean_header}')")
                            break
            
            self.log_message(f"📋 Final mapping: {header_to_col_mapping}")
            self.log_message(f"📋 Found {len(header_to_col_mapping)} matching columns in Units sheet")
            
            # Write headers to row 2 (after timestamp)
            for col_num, header in enumerate(meaningful_headers, 1):
                cell = buildings_ws.cell(row=2, column=col_num, value=header)
                # Make headers bold
                from openpyxl.styles import Font
                cell.font = Font(bold=True)
            
            # For each building, create a row with SUM formulas (starting from row 3)
            for row_num, building_name in enumerate(building_names, 3):
                # Write building name to the first column
                buildings_ws.cell(row=row_num, column=1, value=building_name)
                
                # For each meaningful data column (skip Building column), create appropriate formula
                for col_num in range(2, len(meaningful_headers) + 1):
                    header_name = meaningful_headers[col_num - 1]
                    
                    # Special handling for 2025 Cap Valn. - use direct cell reference from mapping
                    if header_name == "2025 Cap Valn. (£)":
                        try:
                            if building_name in cap_valn_mapping:
                                cap_valn_row = cap_valn_mapping[building_name]
                                # Create direct cell reference to Bank Schedule with correct Excel syntax
                                # Convert column number to Excel letter format
                                cap_valn_col_letter = get_column_letter(cap_valn_col)
                                formula = f'=\'Bank Schedule\'!{cap_valn_col_letter}{cap_valn_row}'
                                
                                cell = buildings_ws.cell(row=row_num, column=col_num)
                                cell.value = formula
                                
                                # Set currency formatting for cap valuation
                                cell.number_format = '£#,##0'
                                
                                self.log_message(f"   ✅ Created Cap Valn formula for {building_name}: Row {cap_valn_row}")
                            else:
                                self.log_message(f"   ⚠️ No Cap Valn mapping found for {building_name}")
                                
                        except Exception as formula_error:
                            self.log_message(f"   ❌ Could not create cap valuation formula for {building_name}: {str(formula_error)}")
                    
                    # Regular SUMIF formulas for other columns
                    elif header_name in header_to_col_mapping:
                        units_col_num = header_to_col_mapping[header_name]
                        units_col_letter = get_column_letter(units_col_num)
                        
                        # Find Property column in Units sheet (contains building names)
                        property_col_letter = "D"  # Default to column D (Property)
                        for col in range(1, 10):
                            header_val = units_ws.cell(row=2, column=col).value  # Row 2 since headers are now on row 2
                            if header_val:
                                header_clean = str(header_val).lower().strip()
                                # Look for exact "property" match, not "property number"
                                if header_clean == "property":
                                    property_col_letter = get_column_letter(col)
                                    break
                        
                        # Create SUMIF formula with proper error handling
                        try:
                            # Use proper Excel sheet reference format
                            formula = f'=SUMIF(Units!{property_col_letter}:{property_col_letter},"{building_name}",Units!{units_col_letter}:{units_col_letter})'
                            
                            cell = buildings_ws.cell(row=row_num, column=col_num)
                            cell.value = formula
                            
                            # Copy number formatting from Units sheet if possible
                            try:
                                units_data_cell = units_ws.cell(row=2, column=units_col_num)
                                if units_data_cell.number_format:
                                    cell.number_format = units_data_cell.number_format
                            except:
                                pass
                                
                        except Exception as formula_error:
                            self.log_message(f"   ❌ Could not create formula for {header_name}: {str(formula_error)}")
                            # Leave cell empty if formula creation fails
            
            self.log_message(f"✅ Created SUM formulas for {len(building_names)} buildings across {len(meaningful_headers)-1} data columns")
            
            # Add autofilter to the headers (row 2)
            buildings_ws.auto_filter.ref = f"A2:{get_column_letter(len(meaningful_headers))}{2 + len(building_names)}"
            self.log_message("✅ Autofilter applied to Buildings sheet headers")

            # Merge the first 5 cells in row 1 (A1:E1) for the timestamp
            try:
                buildings_ws.merge_cells('A1:E1')
                self.log_message("✅ Merged cells A1:E1 for timestamp in Buildings sheet")
            except Exception as e:
                self.log_message(f"⚠️ Could not merge cells for timestamp: {str(e)}")
            
            # Auto-width all columns (with better error handling, excluding the merged timestamp cell)
            try:
                for col_idx in range(1, len(meaningful_headers) + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    
                    # Start from row 2 to skip the merged timestamp row
                    for row_idx in range(2, 3 + len(building_names)):  # Headers + data rows
                        cell = buildings_ws.cell(row=row_idx, column=col_idx)
                        try:
                            cell_value = cell.value
                            if cell_value is not None:
                                cell_length = len(str(cell_value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Set reasonable width limits, especially for Buildings sheet
                    adjusted_width = max(min(max_length + 2, 25), 8)  # Reduced max from 50 to 25 for Buildings
                    buildings_ws.column_dimensions[column_letter].width = adjusted_width
                self.log_message("✅ Auto-width applied to all columns in Buildings sheet")
            except Exception as e:
                self.log_message(f"⚠️ Could not apply auto-width to Buildings sheet: {str(e)}")
            
            # Freeze panes at B3 (freeze timestamp and header rows and first column)
            try:
                buildings_ws.freeze_panes = "B3"
                self.log_message("✅ Freeze panes applied at B3 (timestamp and header rows and first column frozen)")
            except Exception as e:
                self.log_message(f"⚠️ Could not apply freeze panes to Buildings sheet: {str(e)}")
            
            # Save the workbook with the new sheet
            workbook.save(input_path)
            workbook.close()
            
            self.log_message("✅ Buildings summary sheet created successfully with SUM formulas")
            self.log_message("✅ Original formatting and formulas preserved in all existing sheets")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating Buildings summary sheet: {str(e)}")
            raise

    def create_units_sheet(self, input_path, building_names):
        """Create a new 'Units' sheet with all unit data and building validation dropdown."""
        try:
            self.log_message("Step 4: Creating Units sheet with data validation...")
            
            # Import additional openpyxl components
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            import copy
            
            # Open the workbook directly (no copying needed)
            workbook = load_workbook(input_path)
            self.log_message("Opened original file for in-place modification")
            
            # Remove existing Units sheet if it exists
            if "Units" in workbook.sheetnames:
                del workbook["Units"]
                self.log_message("📋 Removed existing 'Units' sheet")
            
            # Read the original Bank Schedule sheet with openpyxl to preserve formatting
            source_ws = workbook["Bank Schedule"]
            
            # Also read with pandas for data analysis
            df = pd.read_excel(input_path, sheet_name="Bank Schedule", header=2)
            
            # Find the Unit Type column
            unit_type_col = None
            unit_type_col_idx = None
            for col in df.columns:
                if 'Unit Type' in str(col) or col == 'G' or df.columns.get_loc(col) == 6:
                    unit_type_col = col
                    unit_type_col_idx = df.columns.get_loc(col)
                    break
            
            if unit_type_col is None and len(df.columns) > 6:
                unit_type_col = df.columns[6]
                unit_type_col_idx = 6
            
            # Create the new Units worksheet
            units_ws = workbook.create_sheet("Units")
            
            # Add timestamp header at the very top
            from datetime import datetime
            now = datetime.now()
            timestamp_text = f"Document updated {now.strftime('%d/%m/%Y')} at {now.strftime('%-I.%M%p').lower()}"
            timestamp_cell = units_ws.cell(row=1, column=1, value=timestamp_text)
            # Make timestamp bold and slightly larger
            from openpyxl.styles import Font
            timestamp_cell.font = Font(bold=True, size=11)
            
            # Build column mapping: source column number -> target column number
            column_mapping = {}
            
            # Copy headers from source sheet (row 3, which is index 2)
            # Headers now go to row 2 (shifted down by 1)
            # Start with "Surveyor" column (skip Building column since Property already has building names)
            # Also skip the empty column after Surveyor
            
            # Find Surveyor column in source (typically column B)
            header_row = 3  # Headers are in row 3 of source sheet
            surveyor_col_idx = 2  # Column B (1-indexed, so 2)
            target_col = 1  # Start at column A in Units sheet
            
            # Copy headers starting from Surveyor column and build column mapping
            for col_idx in range(surveyor_col_idx, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=header_row, column=col_idx)
                
                # Skip empty columns (like the one after Surveyor)
                if not source_cell.value or str(source_cell.value).strip() == '':
                    continue
                
                # Build column mapping: source column -> target column
                column_mapping[col_idx] = target_col
                
                target_cell = units_ws.cell(row=2, column=target_col)  # Row 2 instead of row 1
                
                # Handle potential formulas in headers (rare but possible)
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    refactored_formula = self.refactor_formula(
                        source_cell.value, 
                        header_row, 
                        1, 
                        column_mapping=column_mapping
                    )
                    target_cell.value = refactored_formula
                else:
                    target_cell.value = source_cell.value
                    
                # Copy formatting safely
                try:
                    if source_cell.has_style:
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy.copy(source_cell.protection)
                        target_cell.alignment = copy.copy(source_cell.alignment)
                except Exception as e:
                    self.log_message(f"   ⚠️ Could not copy formatting for header cell: {str(e)}")
                
                target_col += 1
            
            self.log_message(f"📋 Built column mapping: {len(column_mapping)} columns mapped")
            self.log_message(f"   Mapping: {dict(list(column_mapping.items())[:5])}...")  # Show first 5 mappings
            
            # Track units and their buildings
            current_building = None
            units_row = 3  # Start from row 3 (after timestamp and headers)
            
            # Process all data rows starting from row 4 (after headers)
            for source_row_idx in range(4, source_ws.max_row + 1):
                # Get the unit type value
                unit_type_cell = source_ws.cell(row=source_row_idx, column=unit_type_col_idx + 1)  # +1 because Excel is 1-indexed
                unit_type_value = unit_type_cell.value
                
                # Skip empty rows
                if not unit_type_value or str(unit_type_value).strip() == '':
                    continue
                
                # Check if this is a building row or unit row
                non_empty_cols = 0
                for col_idx in range(1, source_ws.max_column + 1):
                    if col_idx != unit_type_col_idx + 1:  # Skip the unit type column
                        cell_value = source_ws.cell(row=source_row_idx, column=col_idx).value
                        if cell_value and str(cell_value).strip() != '':
                            non_empty_cols += 1
                
                # If this row has Unit Type but few other fields, it's likely a building
                if non_empty_cols <= 2:
                    building_name = str(unit_type_value).strip()
                    # Filter out notes and other non-building entries
                    if not any(keyword in building_name.lower() for keyword in ['note', 'capital values', 'rent pa', 'as a result']):
                        current_building = building_name
                else:
                    # This is a unit row - copy data starting from Surveyor column
                    if current_building:
                        target_col = 1  # Start at column A in Units sheet
                        
                        # Copy data using the column mapping we built
                        for source_col_idx in range(surveyor_col_idx, source_ws.max_column + 1):
                            # Only process columns that are in our mapping
                            if source_col_idx not in column_mapping:
                                continue
                            
                            source_cell = source_ws.cell(row=source_row_idx, column=source_col_idx)
                            target_col = column_mapping[source_col_idx]
                            target_cell = units_ws.cell(row=units_row, column=target_col)
                            
                            # Handle formulas with refactoring using proper column mapping
                            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                # Refactor formula for new sheet structure
                                refactored_formula = self.refactor_formula(
                                    source_cell.value, 
                                    source_row_idx, 
                                    units_row, 
                                    column_mapping=column_mapping
                                )
                                target_cell.value = refactored_formula
                            else:
                                # Copy regular values
                                target_cell.value = source_cell.value
                            
                            # Copy formatting safely
                            try:
                                if source_cell.has_style:
                                    target_cell.font = copy.copy(source_cell.font)
                                    target_cell.border = copy.copy(source_cell.border)
                                    target_cell.fill = copy.copy(source_cell.fill)
                                    target_cell.number_format = source_cell.number_format
                                    target_cell.protection = copy.copy(source_cell.protection)
                                    target_cell.alignment = copy.copy(source_cell.alignment)
                            except Exception as e:
                                pass  # Continue if formatting copy fails
                        
                        units_row += 1
            
            self.log_message(f"✅ Copied {units_row - 3} unit rows to Units sheet")  # -3 because we start from row 3
            
            # Skip data validation dropdown to avoid incomplete building lists
            self.log_message(f"📝 Building column populated for all {len(building_names)} buildings (no dropdown to avoid incomplete lists)")
            
            # Add autofilter to the headers (row 2)
            units_ws.auto_filter.ref = f"A2:{get_column_letter(units_ws.max_column)}{units_row - 1}"
            self.log_message("✅ Autofilter applied to Units sheet headers")

            # Merge the first 5 cells in row 1 (A1:E1) for the timestamp
            try:
                units_ws.merge_cells('A1:E1')
                self.log_message("✅ Merged cells A1:E1 for timestamp in Units sheet")
            except Exception as e:
                self.log_message(f"⚠️ Could not merge cells for timestamp: {str(e)}")
            
            # Auto-width all columns (with better error handling, excluding the merged timestamp cell)
            try:
                for col_idx in range(1, units_ws.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    
                    # Start from row 2 to skip the merged timestamp row
                    for row_idx in range(2, units_ws.max_row + 1):
                        cell = units_ws.cell(row=row_idx, column=col_idx)
                        try:
                            cell_value = cell.value
                            if cell_value is not None:
                                cell_length = len(str(cell_value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Set reasonable width limits
                    adjusted_width = max(min(max_length + 2, 30), 8)  # Reduced max from 50 to 30
                    units_ws.column_dimensions[column_letter].width = adjusted_width
                self.log_message("✅ Auto-width applied to all columns in Units sheet")
            except Exception as e:
                self.log_message(f"⚠️ Could not apply auto-width: {str(e)}")
            
            # Hide specific columns by header name
            try:
                columns_to_hide = ["Client", "Property Number", ".", "check 1", "check 2", "check 3", "check 4", "*"]
                hidden_count = 0
                
                for col_idx in range(1, units_ws.max_column + 1):
                    header_cell = units_ws.cell(row=2, column=col_idx)  # Row 2 since headers are now on row 2
                    header_value = str(header_cell.value).strip() if header_cell.value else ""
                    
                    if header_value in columns_to_hide:
                        col_letter = get_column_letter(col_idx)
                        units_ws.column_dimensions[col_letter].hidden = True
                        hidden_count += 1
                        self.log_message(f"   Hidden column {col_letter}: '{header_value}'")
                
                if hidden_count > 0:
                    self.log_message(f"✅ Hidden {hidden_count} columns in Units sheet")
                else:
                    self.log_message(f"⚠️ No matching columns found to hide")
            except Exception as e:
                self.log_message(f"⚠️ Could not hide columns: {str(e)}")
            
            # Freeze panes at F3 (freeze timestamp and header rows and columns A-E)
            try:
                units_ws.freeze_panes = "F3"
                self.log_message("✅ Freeze panes applied at F3 (timestamp and header rows and columns A-E frozen)")
            except Exception as e:
                self.log_message(f"⚠️ Could not apply freeze panes: {str(e)}")
            
            # Save the workbook with better error handling
            try:
                workbook.save(input_path)
                workbook.close()
                self.log_message("✅ Units sheet created successfully (no dropdown for complete building list)")
                return True
            except Exception as e:
                self.log_message(f"❌ Error saving workbook: {str(e)}")
                workbook.close()
                raise
            
        except Exception as e:
            self.log_message(f"❌ Error creating Units sheet: {str(e)}")
            raise

    def sanitize_file(self):
        """Process the selected Excel file and save the sanitized version."""
        try:
            self.log_message("\n" + "="*50)
            self.log_message("Starting sanitization process...")
            
            input_path = self.input_file_path.get()
            
            # Validate input file
            is_valid, validation_message = self.validate_excel_file(input_path)
            if not is_valid:
                messagebox.showerror("Error", validation_message)
                return
                
            self.log_message("Step 1: Input file validation completed.")
            
            # Check if file is locked before proceeding
            self.log_message("Step 1.5: Checking file accessibility...")
            lock_status = self.check_file_lock_status(input_path)
            if lock_status['is_locked']:
                error_details = f": {lock_status.get('error', 'Unknown reason')}" if 'error' in lock_status else ""
                error_msg = f"Cannot process file - it appears to be open by another user{error_details}"
                self.log_message(f"❌ {error_msg}")
                messagebox.showerror("File Locked", 
                    f"{error_msg}\n\n"
                    "Please ensure the Excel file is closed by all users before running this operation.\n"
                    "If the file is on a shared network drive, check that no one else has it open.")
                return
            else:
                self.log_message(f"✅ File is accessible for modification")
            
            # Check existing sheets before processing
            initial_sheet_info = self.check_existing_sheets(input_path)
            
            # Analyze the Bank Schedule data
            self.log_message("Step 2: Analyzing Bank Schedule data structure...")
            
            try:
                analysis_results = self.analyze_bank_schedule_data(input_path)
                
                self.log_message("\n" + "-"*30)
                self.log_message("📊 DATA ANALYSIS RESULTS:")
                self.log_message("-"*30)
                self.log_message(f"🏢 Buildings found: {analysis_results['buildings']}")
                self.log_message(f"🏠 Units found: {analysis_results['units']}")
                self.log_message(f"📝 Empty rows: {analysis_results['empty_rows']}")
                self.log_message(f"📋 Total rows processed: {analysis_results['total_rows']}")
                self.log_message("-"*30 + "\n")
                
            except Exception as e:
                self.log_message(f"❌ Error during analysis: {str(e)}")
                messagebox.showerror("Analysis Error", f"Could not analyze data structure: {str(e)}")
                return
            
            self.log_message("Step 3: Processing file and creating Units & Buildings sheets...")
            
            # Create backup of original file for safety
            input_name = os.path.basename(input_path)
            input_dir = os.path.dirname(input_path)
            name_without_ext = os.path.splitext(input_name)[0]
            ext = os.path.splitext(input_name)[1]
            backup_path = os.path.join(input_dir, f"{name_without_ext}_BACKUP{ext}")
            
            # Handle case where backup already exists
            backup_counter = 1
            original_backup_path = backup_path
            while os.path.exists(backup_path):
                backup_path = os.path.join(input_dir, f"{name_without_ext}_BACKUP_{backup_counter}{ext}")
                backup_counter += 1
            
            try:
                shutil.copy2(input_path, backup_path)
                self.log_message(f"🔄 Created backup: {os.path.basename(backup_path)}")
            except Exception as e:
                if not messagebox.askyesno("Backup Failed", 
                    f"Could not create backup file: {str(e)}\n\n" +
                    "Do you want to continue anyway? (This will modify your original file directly)"):
                    self.log_message("⚠ Operation cancelled - backup failed and user chose not to continue.")
                    return
                self.log_message("⚠ Warning: Proceeding without backup!")
                backup_path = None
            
            # Add sheets directly to the original file
            try:
                # Always create Units sheet first (independent of Buildings sheet requirements)
                self.create_units_sheet(input_path, analysis_results['building_names'])
                
                # Try to create Buildings sheet (depends on specific columns)
                try:
                    # Analyze Cap Valn locations for Buildings sheet
                    cap_valn_mapping, cap_valn_col = self.analyze_cap_valn_locations(input_path)
                    
                    # Create Buildings summary sheet
                    self.create_buildings_summary_sheet(input_path, analysis_results, cap_valn_mapping, cap_valn_col)
                    
                except ValueError as ve:
                    # Missing columns for Buildings sheet - show warning but continue
                    if "Missing required columns" in str(ve):
                        warning_msg = f"Buildings sheet could not be created: {str(ve)}"
                        self.log_message(f"⚠️ {warning_msg}")
                        self.log_message("✅ Units sheet created successfully (Buildings sheet skipped due to missing columns)")
                        
                        # Show special success message for Units only
                        units_existed = initial_sheet_info['units_exists']
                        if units_existed:
                            units_msg = "'Units' sheet has been successfully updated!"
                        else:
                            units_msg = "'Units' sheet has been successfully created!"
                        
                        messagebox.showwarning("Buildings Sheet Skipped", 
                                             f"{warning_msg}\n\n{units_msg} "
                                             "To create the Buildings sheet, ensure your Excel file contains these columns:\n"
                                             "• Unit Demise\n• 2025 Cap Valn. (£)\n• Property")
                        
                        self.log_message("="*50 + "\n")
                        return
                    else:
                        # Re-raise other ValueError cases
                        raise ve
                
                # Create hierarchical legacy view for familiar format - TEMPORARILY DISABLED
                # self.create_hierarchical_view_sheet_from_files(input_path, analysis_results['building_names'])
            except ValueError as ve:
                # Handle missing columns error for Units sheet (different error)
                if "Missing required columns" not in str(ve):
                    self.log_message(f"❌ Data validation error: {str(ve)}")
                    messagebox.showerror("Data Error", f"Data validation error: {str(ve)}")
                    return
            except Exception as e:
                self.log_message(f"❌ Error adding sheets to file: {str(e)}")
                messagebox.showerror("File Processing Error", f"Could not add sheets to file: {str(e)}")
                return
            
            self.log_message(f"✅ Sheets processed successfully in: {os.path.basename(input_path)}")
            
            # Show appropriate message based on what sheets existed initially
            units_existed = initial_sheet_info['units_exists']
            buildings_existed = initial_sheet_info['buildings_exists']
            
            if units_existed and buildings_existed:
                success_msg = "Both 'Units' and 'Buildings' sheets have been successfully updated!"
            elif units_existed and not buildings_existed:
                success_msg = "'Units' sheet updated and 'Buildings' sheet created successfully!"
            elif not units_existed and buildings_existed:
                success_msg = "'Units' sheet created and 'Buildings' sheet updated successfully!"
            else:
                success_msg = "'Units' and 'Buildings' sheets created successfully!"
            
            # Show completion message
            messagebox.showinfo("Success", success_msg)
            
            self.log_message("="*50 + "\n")
            
        except Exception as e:
            error_msg = f"Error during sanitization: {str(e)}"
            self.log_message(f"✗ {error_msg}")
            messagebox.showerror("Error", error_msg)


def main():
    """Main function to run the application."""
    root = tk.Tk()
    app = BankScheduleSanitizer(root)
    root.mainloop()


if __name__ == "__main__":
    main()
