#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook

# Test different Excel sheet reference formats
wb = Workbook()

# Create a Units sheet with test data
units_ws = wb.active
units_ws.title = "Units"
units_ws.cell(row=1, column=1, value="Building")
units_ws.cell(row=1, column=2, value="Test Data")
units_ws.cell(row=2, column=1, value="Building A")
units_ws.cell(row=2, column=2, value="Test Value")

# Create Legacy View sheet
legacy_ws = wb.create_sheet("Legacy View")

# Test different formula formats
legacy_ws.cell(row=1, column=1, value="Format Test")
legacy_ws.cell(row=2, column=1, value="Format 1:")
legacy_ws.cell(row=2, column=2, value="=Units.A2")  # Original format
legacy_ws.cell(row=3, column=1, value="Format 2:")
legacy_ws.cell(row=3, column=2, value="='Units'.A2")  # With quotes
legacy_ws.cell(row=4, column=1, value="Format 3:")
legacy_ws.cell(row=4, column=2, value="=Units!A2")  # Excel standard format
legacy_ws.cell(row=5, column=1, value="Format 4:")
legacy_ws.cell(row=5, column=2, value="='Units'!A2")  # Excel standard with quotes

# Save test file
test_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/formula_test.xlsx"
wb.save(test_path)
print(f"Test file saved: {test_path}")
print("Please open this file in Excel to see which formula format works correctly.")
