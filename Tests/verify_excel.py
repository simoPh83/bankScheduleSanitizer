#!/usr/bin/env python3
"""
Simple verification script to check if our sheets are created correctly
"""

from openpyxl import load_workbook
import sys
import os

def verify_excel_file(file_path):
    """Verify the Excel file has the expected sheets and features."""
    
    try:
        workbook = load_workbook(file_path)
        
        print(f"📋 Sheets found: {workbook.sheetnames}")
        
        # Check Buildings sheet
        if 'Buildings' in workbook.sheetnames:
            buildings_ws = workbook['Buildings']
            print(f"✅ Buildings sheet: {buildings_ws.max_row} rows, {buildings_ws.max_column} columns")
            print(f"   🎛️ Autofilter: {buildings_ws.auto_filter.ref}")
        
        # Check Units sheet  
        if 'Units' in workbook.sheetnames:
            units_ws = workbook['Units']
            print(f"✅ Units sheet: {units_ws.max_row} rows, {units_ws.max_column} columns")
            print(f"   🎛️ Autofilter: {units_ws.auto_filter.ref}")
            
            # Check data validation
            validations = list(units_ws.data_validations)
            if validations:
                print(f"   ✅ Data validation applied to: {validations[0].sqref}")
            else:
                print("   ❌ No data validation found")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python verify_excel.py <excel_file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    print(f"🔍 Verifying Excel file: {os.path.basename(file_path)}")
    
    success = verify_excel_file(file_path)
    sys.exit(0 if success else 1)
