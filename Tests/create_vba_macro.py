#!/usr/bin/env python3

import openpyxl
from openpyxl import load_workbook
import os
import shutil

def create_vba_macro_file():
    """Create a template macro-enabled Excel file with VBA code for dynamic Legacy View"""
    
    # VBA macro code as a text file that can be imported
    vba_code = '''Sub RefreshLegacyView()
    ' Dynamic Legacy View Rebuilder
    ' Parses Units sheet and recreates hierarchical Legacy View structure
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    
    On Error GoTo ErrorHandler
    
    Dim unitsWs As Worksheet
    Dim legacyWs As Worksheet
    Dim lastRow As Long
    Dim currentRow As Long
    Dim headerRow As Long
    Dim col As Long
    Dim buildingDict As Object
    Dim buildingName As String
    Dim key As Variant
    Dim unitRows As Collection
    Dim i As Long
    Dim startRow As Long, endRow As Long
    
    ' Get worksheets
    Set unitsWs = ThisWorkbook.Worksheets("Units")
    
    ' Clear existing Legacy View content (keep first 3 instruction rows)
    Set legacyWs = ThisWorkbook.Worksheets("Legacy View")
    lastRow = legacyWs.Cells(legacyWs.Rows.Count, 1).End(xlUp).Row
    If lastRow > 3 Then
        legacyWs.Range("A4:ZZ" & lastRow).Clear
    End If
    
    ' Set up headers
    headerRow = 4
    For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
        If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
            legacyWs.Cells(headerRow, col - 1).Value = unitsWs.Cells(1, col).Value
            legacyWs.Cells(headerRow, col - 1).Font.Bold = True
        End If
    Next col
    
    ' Group units by building
    Set buildingDict = CreateObject("Scripting.Dictionary")
    lastRow = unitsWs.Cells(unitsWs.Rows.Count, 1).End(xlUp).Row
    
    For i = 2 To lastRow
        buildingName = Trim(CStr(unitsWs.Cells(i, 1).Value))
        If buildingName <> "" Then
            If Not buildingDict.Exists(buildingName) Then
                Set buildingDict(buildingName) = New Collection
            End If
            buildingDict(buildingName).Add i
        End If
    Next i
    
    ' Build hierarchical structure
    currentRow = headerRow + 2 ' Empty row after headers
    
    For Each key In buildingDict.Keys
        Set unitRows = buildingDict(key)
        
        ' Building header
        legacyWs.Cells(currentRow, 1).Value = key
        legacyWs.Cells(currentRow, 1).Font.Bold = True
        legacyWs.Cells(currentRow, 1).Font.Size = 12
        currentRow = currentRow + 2 ' Empty row after building header
        
        ' Unit rows
        startRow = currentRow
        For i = 1 To unitRows.Count
            For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
                If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
                    legacyWs.Cells(currentRow, col - 1).Formula = "=Units!" & _
                        ConvertToColumnLetter(col) & unitRows(i)
                End If
            Next col
            currentRow = currentRow + 1
        Next i
        endRow = currentRow - 1
        
        ' Empty row + summary row
        currentRow = currentRow + 1
        legacyWs.Cells(currentRow, 1).Value = key & " - TOTAL"
        legacyWs.Cells(currentRow, 1).Font.Bold = True
        legacyWs.Cells(currentRow, 1).Font.Italic = True
        
        ' Add SUM formulas for numeric columns
        For col = 2 To unitsWs.Cells(1, unitsWs.Columns.Count).End(xlToLeft).Column
            If Not IsEmpty(unitsWs.Cells(1, col).Value) Then
                If IsNumericColumn(unitsWs.Cells(1, col).Value) Then
                    If startRow <= endRow Then
                        legacyWs.Cells(currentRow, col - 1).Formula = "=SUM(" & _
                            ConvertToColumnLetter(col - 1) & startRow & ":" & _
                            ConvertToColumnLetter(col - 1) & endRow & ")"
                        legacyWs.Cells(currentRow, col - 1).Font.Bold = True
                    End If
                End If
            End If
        Next col
        
        currentRow = currentRow + 2 ' Empty row after summary
    Next key
    
    ' Auto-fit columns
    legacyWs.Columns.AutoFit
    
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    
    MsgBox "Legacy View refreshed successfully! Found " & buildingDict.Count & " buildings.", vbInformation
    Exit Sub
    
ErrorHandler:
    Application.Calculation = xlCalculationAutomatic
    Application.ScreenUpdating = True
    MsgBox "Error refreshing Legacy View: " & Err.Description, vbCritical
End Sub

Private Function ConvertToColumnLetter(colNum As Long) As String
    Dim result As String
    result = ""
    While colNum > 0
        colNum = colNum - 1
        result = Chr(65 + (colNum Mod 26)) & result
        colNum = colNum \\ 26
    Wend
    ConvertToColumnLetter = result
End Function

Private Function IsNumericColumn(header As String) As Boolean
    Dim lowerHeader As String
    lowerHeader = LCase(Trim(header))
    
    IsNumericColumn = (InStr(lowerHeader, "area") > 0) Or _
                      (InStr(lowerHeader, "rent") > 0) Or _
                      (InStr(lowerHeader, "erv") > 0) Or _
                      (InStr(lowerHeader, "value") > 0) Or _
                      (InStr(lowerHeader, "valuation") > 0) Or _
                      (InStr(lowerHeader, "cap") > 0) Or _
                      (InStr(lowerHeader, "£") > 0) Or _
                      (InStr(lowerHeader, "$") > 0) Or _
                      (InStr(lowerHeader, "sq") > 0) Or _
                      (InStr(lowerHeader, "total") > 0) Or _
                      (InStr(lowerHeader, "sum") > 0) Or _
                      (InStr(lowerHeader, "cost") > 0)
End Function

Sub Auto_Open()
    ' Set up keyboard shortcut when workbook opens
    Application.OnKey "^+R", "RefreshLegacyView"
End Sub'''
    
    # Save VBA code to a file for manual import
    vba_file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/RefreshLegacyView_VBA.bas"
    
    with open(vba_file_path, 'w') as f:
        f.write(vba_code)
    
    print(f"✅ VBA macro code saved to: {vba_file_path}")
    print("\n📋 INSTRUCTIONS TO ADD VBA MACRO:")
    print("1. Open your Excel file in Excel")
    print("2. Press Alt+F11 to open VBA Editor")
    print("3. Right-click on your workbook in Project Explorer")
    print("4. Choose Insert > Module")
    print("5. Copy-paste the VBA code from the .bas file")
    print("6. Save as .xlsm (macro-enabled) format")
    print("7. Use Ctrl+Shift+R to refresh Legacy View anytime")
    print("\n🔧 ALTERNATIVE: Use Developer > Import File to import the .bas file directly")

def add_vba_instructions_to_excel(file_path):
    """Add detailed VBA instructions to the Legacy View sheet"""
    try:
        print(f"Adding VBA instructions to: {file_path}")
        
        workbook = load_workbook(file_path)
        
        if "Legacy View" in workbook.sheetnames:
            legacy_ws = workbook["Legacy View"]
            
            # Update instructions
            legacy_ws.cell(row=1, column=1, value="Legacy View - Dynamic Hierarchical Structure")
            legacy_ws.cell(row=2, column=1, value="(Responsive to Units sheet changes - requires VBA macro)")
            legacy_ws.cell(row=3, column=1, value="SETUP: Import VBA macro from RefreshLegacyView_VBA.bas file")
            
            # Add detailed instructions in column E
            instructions = [
                "VBA MACRO SETUP INSTRUCTIONS:",
                "",
                "1. Press Alt+F11 to open VBA Editor",
                "2. Right-click workbook in Project Explorer",
                "3. Choose Insert > Module",
                "4. Copy VBA code from RefreshLegacyView_VBA.bas",
                "5. Save file as .xlsm format",
                "6. Press Ctrl+Shift+R to refresh anytime",
                "",
                "BENEFITS:",
                "• Handles any number of buildings/units",
                "• Auto-rebuilds structure when data changes",
                "• Preserves hierarchical format",
                "• Automatic SUM formulas for totals",
                "",
                "USAGE:",
                "• Run macro after Units sheet changes",
                "• Keyboard shortcut: Ctrl+Shift+R",
                "• Or use Developer > Macros > RefreshLegacyView"
            ]
            
            for i, instruction in enumerate(instructions):
                legacy_ws.cell(row=i + 1, column=5, value=instruction)
                if instruction.endswith(":"):
                    cell = legacy_ws.cell(row=i + 1, column=5)
                    if hasattr(openpyxl.styles, 'Font'):
                        cell.font = openpyxl.styles.Font(bold=True)
            
            workbook.save(file_path)
            workbook.close()
            print("✅ VBA instructions added to Legacy View sheet")
            
    except Exception as e:
        print(f"⚠️ Error adding instructions: {str(e)}")

if __name__ == "__main__":
    # Create VBA file
    create_vba_macro_file()
    
    # Add instructions to existing Excel file
    excel_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/8.xlsx"
    if os.path.exists(excel_file):
        add_vba_instructions_to_excel(excel_file)
