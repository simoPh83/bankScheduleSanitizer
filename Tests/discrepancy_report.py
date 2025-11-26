#!/usr/bin/env python3
import openpyxl

def generate_discrepancy_report():
    """Generate a comprehensive report of building name discrepancies between Buildings sheet and Bank Schedule"""
    
    print("🔍 BUILDING NAME DISCREPANCY REPORT")
    print("=" * 80)
    print()
    
    try:
        # Get all buildings from Buildings sheet
        buildings_wb = openpyxl.load_workbook('data/27.xlsx', data_only=False)
        buildings_ws = buildings_wb['Buildings']
        
        buildings_list = []
        cap_valn_status = {}
        
        # Find Cap Valn column
        cap_valn_col = None
        for col in range(1, 10):
            header_val = buildings_ws.cell(row=1, column=col).value
            if header_val and "2024 Cap Valn" in str(header_val):
                cap_valn_col = col
                break
        
        for row in range(2, 100):  # Check up to row 100
            building_name = buildings_ws.cell(row=row, column=1).value
            if building_name:
                building_name = str(building_name).strip()
                buildings_list.append(building_name)
                
                # Check if building has Cap Valn formula
                if cap_valn_col:
                    formula = buildings_ws.cell(row=row, column=cap_valn_col).value
                    cap_valn_status[building_name] = {
                        'row': row,
                        'has_formula': formula is not None,
                        'formula': formula
                    }
            else:
                break
        
        buildings_wb.close()
        print(f"📋 Found {len(buildings_list)} buildings in Buildings sheet")
        
        # Get all buildings from Bank Schedule
        bank_wb = openpyxl.load_workbook('data/Leasing Bank Schedule June 2025.xlsx', data_only=False)
        bank_ws = bank_wb['Bank Schedule']
        
        bank_buildings = set()
        bank_building_details = {}
        
        for row in range(1, 1000):  # Check more rows
            property_val = bank_ws.cell(row=row, column=5).value  # Column E (Property)
            if property_val and str(property_val).strip():
                building_name = str(property_val).strip()
                bank_buildings.add(building_name)
                
                if building_name not in bank_building_details:
                    unit_demise = bank_ws.cell(row=row, column=6).value  # Column F (Unit Demise)
                    cap_valn = bank_ws.cell(row=row, column=28).value  # Column AB (Cap Valn)
                    bank_building_details[building_name] = {
                        'first_row': row,
                        'has_unit_demise': bool(unit_demise and str(unit_demise).strip()),
                        'has_cap_valn': bool(cap_valn and str(cap_valn).strip() and str(cap_valn).strip() != 'None'),
                        'sample_cap_valn': cap_valn
                    }
        
        bank_wb.close()
        print(f"📋 Found {len(bank_buildings)} unique building names in Bank Schedule")
        print()
        
        # Analysis 1: Buildings in Buildings sheet but not in Bank Schedule (exact match)
        print("❌ BUILDINGS IN BUILDINGS SHEET BUT NOT FOUND IN BANK SCHEDULE (EXACT MATCH):")
        print("-" * 80)
        missing_exact = []
        for building in buildings_list:
            if building not in bank_buildings:
                status = cap_valn_status.get(building, {})
                missing_exact.append({
                    'name': building,
                    'row': status.get('row'),
                    'has_formula': status.get('has_formula', False),
                    'formula': status.get('formula')
                })
        
        if missing_exact:
            for i, building in enumerate(missing_exact, 1):
                print(f"{i:2d}. Row {building['row']:2d}: '{building['name']}'")
                print(f"     Has Cap Valn Formula: {building['has_formula']}")
                if building['formula']:
                    print(f"     Formula: {building['formula']}")
                print()
        else:
            print("✅ All buildings found with exact name match")
        
        print(f"Total: {len(missing_exact)} buildings missing")
        print()
        
        # Analysis 2: Find potential fuzzy matches for missing buildings
        print("🔍 POTENTIAL FUZZY MATCHES FOR MISSING BUILDINGS:")
        print("-" * 80)
        
        def find_fuzzy_matches(target, candidates, threshold=0.6):
            """Simple fuzzy matching based on common words"""
            target_words = set(target.lower().split())
            matches = []
            
            for candidate in candidates:
                candidate_words = set(candidate.lower().split())
                # Calculate Jaccard similarity (intersection over union)
                intersection = len(target_words.intersection(candidate_words))
                union = len(target_words.union(candidate_words))
                similarity = intersection / union if union > 0 else 0
                
                if similarity >= threshold:
                    matches.append((candidate, similarity))
            
            return sorted(matches, key=lambda x: x[1], reverse=True)
        
        for building in missing_exact:
            building_name = building['name']
            matches = find_fuzzy_matches(building_name, bank_buildings, threshold=0.3)
            
            print(f"🎯 '{building_name}' (Row {building['row']}):")
            if matches:
                print("   Potential matches in Bank Schedule:")
                for match, similarity in matches[:5]:  # Top 5 matches
                    bank_info = bank_building_details[match]
                    print(f"   • '{match}' (similarity: {similarity:.2f})")
                    print(f"     Row {bank_info['first_row']}, Has Unit Demise: {bank_info['has_unit_demise']}, Has Cap Valn: {bank_info['has_cap_valn']}")
            else:
                print("   ❌ No potential matches found")
            print()
        
        # Analysis 3: Buildings with missing Cap Valn formulas
        print("📊 SUMMARY OF CAP VALN FORMULA STATUS:")
        print("-" * 80)
        
        with_formula = [b for b in buildings_list if cap_valn_status[b]['has_formula']]
        without_formula = [b for b in buildings_list if not cap_valn_status[b]['has_formula']]
        
        print(f"✅ Buildings with Cap Valn formulas: {len(with_formula)}")
        print(f"❌ Buildings without Cap Valn formulas: {len(without_formula)}")
        print()
        
        if without_formula:
            print("Buildings without Cap Valn formulas:")
            for building in without_formula:
                info = cap_valn_status[building]
                print(f"  • Row {info['row']:2d}: {building}")
            print()
        
        # Analysis 4: Check for buildings that might be aggregated differently
        print("🔄 BUILDINGS THAT MIGHT HAVE NAMING VARIATIONS:")
        print("-" * 80)
        
        # Look for patterns like ranges, abbreviations, etc.
        naming_patterns = []
        for building in missing_exact:
            building_name = building['name']
            # Check for common variations
            potential_variations = []
            
            # Check if it's a range that might be split
            if ' and ' in building_name or ' & ' in building_name:
                potential_variations.append("Might be split into separate buildings")
            
            # Check if it contains numbers that might be ranges
            import re
            if re.search(r'\d+-\d+', building_name):
                potential_variations.append("Contains number range")
            
            # Check for common abbreviations
            abbrev_map = {
                'Street': ['St', 'Str'],
                'House': ['Hse', 'Ho'],
                'Building': ['Bldg', 'Bld'],
                'Floor': ['Fl', 'Flr']
            }
            
            for full, abbrevs in abbrev_map.items():
                if full in building_name:
                    for abbrev in abbrevs:
                        potential_variations.append(f"'{full}' might be abbreviated as '{abbrev}'")
            
            if potential_variations:
                naming_patterns.append({
                    'name': building_name,
                    'variations': potential_variations
                })
        
        if naming_patterns:
            for pattern in naming_patterns:
                print(f"🔄 '{pattern['name']}':")
                for variation in pattern['variations']:
                    print(f"   • {variation}")
                print()
        else:
            print("✅ No obvious naming pattern issues detected")
        
    except Exception as e:
        print(f"❌ Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_discrepancy_report()
