#!/usr/bin/env python3
import openpyxl

def check_specific_building():
    """Check for 50 Eastcastle Street and other buildings with missing Cap Valn"""
    
    try:
        # Load the latest generated file
        wb = openpyxl.load_workbook('data/27.xlsx', data_only=False)
        
        if 'Buildings' not in wb.sheetnames:
            print("❌ Buildings sheet not found")
            return
            
        buildings_ws = wb['Buildings']
        
        print("📋 All Buildings in Buildings Sheet:")
        print("=" * 80)
        
        # Find the Cap Valn column
        cap_valn_col = None
        for col in range(1, 10):
            header_val = buildings_ws.cell(row=1, column=col).value
            if header_val and "2024 Cap Valn" in str(header_val):
                cap_valn_col = col
                break
        
        if not cap_valn_col:
            print("❌ Cap Valn column not found")
            return
        
        # Check all buildings
        buildings_with_no_formula = []
        buildings_with_eastcastle = []
        
        for row in range(2, 100):  # Check more rows to find all buildings
            building_name = buildings_ws.cell(row=row, column=1).value
            if building_name:
                building_name = str(building_name).strip()
                formula_cell = buildings_ws.cell(row=row, column=cap_valn_col)
                formula = formula_cell.value
                
                # Check if this building contains "Eastcastle"
                if "eastcastle" in building_name.lower():
                    buildings_with_eastcastle.append({
                        'row': row,
                        'name': building_name,
                        'formula': formula
                    })
                
                # Check for missing formulas
                if not formula:
                    buildings_with_no_formula.append({
                        'row': row,
                        'name': building_name
                    })
                
                print(f"Row {row:2d}: {building_name}")
                print(f"         Formula: {formula}")
                print()
            else:
                print(f"Row {row:2d}: No building name - stopping")
                break
        
        print("\n🔍 EASTCASTLE BUILDINGS:")
        print("=" * 50)
        for building in buildings_with_eastcastle:
            print(f"Row {building['row']:2d}: {building['name']}")
            print(f"         Formula: {building['formula']}")
            print()
        
        print(f"\n❌ BUILDINGS WITHOUT CAP VALN FORMULAS ({len(buildings_with_no_formula)}):")
        print("=" * 60)
        for building in buildings_with_no_formula:
            print(f"Row {building['row']:2d}: {building['name']}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error checking buildings: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_specific_building()
