#!/usr/bin/env python3
"""
Bank Schedule Sanitizer
A GUI application to sanitize Excel bank schedule files.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import pandas as pd
import shutil
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy
import re


class BankScheduleSanitizer:
    def __init__(self, root):
        self.root = root
        self.root.title("Bank Schedule Sanitizer")
        self.root.geometry("600x500")
        
        # Variables
        self.input_file_path = tk.StringVar()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Set up the user interface."""
        # Main frame
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = tk.Label(
            main_frame, 
            text="Bank Schedule Sanitizer", 
            font=("Arial", 16, "bold")
        )
        title_label.pack(pady=(0, 20))
        
        # Input file selection
        input_frame = tk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(input_frame, text="Select Excel file to process:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        input_path_frame = tk.Frame(input_frame)
        input_path_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.input_path_entry = tk.Entry(
            input_path_frame, 
            textvariable=self.input_file_path, 
            state="readonly",
            font=("Arial", 9)
        )
        self.input_path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        tk.Button(
            input_path_frame, 
            text="Browse...", 
            command=self.browse_input_file,
            width=10
        ).pack(side=tk.RIGHT, padx=(10, 0))
        
        # Sanitize button
        self.sanitize_button = tk.Button(
            main_frame, 
            text="Sanitize", 
            command=self.sanitize_file,
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            height=2,
            state=tk.DISABLED
        )
        self.sanitize_button.pack(pady=20)
        
        # Error/Status text box
        status_frame = tk.Frame(main_frame)
        status_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        tk.Label(status_frame, text="Status and Error Messages:", font=("Arial", 10, "bold")).pack(anchor=tk.W)
        
        self.status_text = scrolledtext.ScrolledText(
            status_frame,
            height=8,
            font=("Consolas", 9),
            wrap=tk.WORD
        )
        self.status_text.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
        
        # Initial status message
        self.log_message("Ready to process Excel files. Please select an input file to begin.")
        
    def browse_input_file(self):
        """Open file dialog to select input Excel file."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File to Process",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.input_file_path.set(file_path)
            self.log_message(f"Selected input file: {os.path.basename(file_path)}")
            # Enable sanitize button when input file is selected
            self.sanitize_button.config(state=tk.NORMAL)
            
    def log_message(self, message):
        """Add a message to the status text box."""
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
        
    def validate_excel_file(self, file_path):
        """Validate that the Excel file exists and can be read."""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
                
            # Try to read the Excel file to check if it's valid
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            self.log_message(f"Excel file validated. Found sheets: {', '.join(sheet_names)}")
            
            # Check if "Bank Schedule" sheet exists
            if "Bank Schedule" in sheet_names:
                self.log_message("✓ Found 'Bank Schedule' sheet in the file.")
                return True, "File is valid"
            else:
                self.log_message("⚠ Warning: 'Bank Schedule' sheet not found in the file.")
                return True, "File is valid but 'Bank Schedule' sheet not found"
                
        except Exception as e:
            error_msg = f"Error validating Excel file: {str(e)}"
            self.log_message(f"✗ {error_msg}")
            return False, error_msg
            
    def refactor_formula(self, formula, source_row, target_row, column_offset=1):
        """
        Refactor Excel formula by adjusting cell references for new sheet structure.
        
        Args:
            formula: The original formula string (e.g., '=V7/H7')
            source_row: Original row number in source sheet
            target_row: Target row number in new sheet
            column_offset: Column shift due to added columns (default 1 for Building column)
        
        Returns:
            Refactored formula string
        """
        import re
        
        if not formula or not formula.startswith('='):
            return formula
        
        try:
            # Pattern to match cell references like H7, V7, etc.
            cell_pattern = r'([A-Z]+)(\d+)'
            
            def replace_cell_ref(match):
                col_letters = match.group(1)
                old_row = int(match.group(2))
                
                # Convert column letters to column number
                old_col_num = 0
                for char in col_letters:
                    old_col_num = old_col_num * 26 + (ord(char) - ord('A') + 1)
                
                # Apply column shift (add offset for Building column)
                new_col_num = old_col_num + column_offset
                
                # Calculate new row number
                # If the original reference was to the same row as source, map it to target row
                if old_row == source_row:
                    new_row = target_row
                else:
                    # For references to other rows, maintain relative offset
                    row_diff = old_row - source_row
                    new_row = target_row + row_diff
                    
                # Ensure row number is positive
                if new_row < 1:
                    new_row = 1
                
                # Convert new column number back to letters
                new_col_letters = ""
                temp_col = new_col_num
                while temp_col > 0:
                    temp_col -= 1
                    new_col_letters = chr(temp_col % 26 + ord('A')) + new_col_letters
                    temp_col //= 26
                
                return f"{new_col_letters}{new_row}"
            
            # Replace all cell references in the formula
            refactored_formula = re.sub(cell_pattern, replace_cell_ref, formula)
            
            self.log_message(f"   📐 Formula refactored: {formula} → {refactored_formula}")
            return refactored_formula
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not refactor formula {formula}: {str(e)}")
            return formula  # Return original if refactoring fails

    def create_hierarchical_view_sheet(self, workbook, units_ws, units_row):
        """Create a dynamic Legacy View sheet with formulas that auto-update from Units sheet"""
        try:
            self.log_message("📋 Creating dynamic Legacy View sheet...")
            
            # Create the legacy view sheet (remove existing ones if they exist)
            if "Hierarchical View" in workbook.sheetnames:
                del workbook["Hierarchical View"]
            if "Hierarchical View1" in workbook.sheetnames:
                del workbook["Hierarchical View1"]
            if "Legacy View" in workbook.sheetnames:
                del workbook["Legacy View"]
                
            hier_ws = workbook.create_sheet("Legacy View")
            
            # Add title and instructions
            hier_ws.cell(row=1, column=1, value="Legacy View - Dynamic Building Structure")
            hier_ws.cell(row=2, column=1, value="(Auto-updates from Units sheet - formulas link to live data)")
            
            # Copy headers from Units sheet starting at row 4
            headers_row = 4
            
            # First, set up column mapping and headers
            col_mapping = {}  # source_col -> target_col
            target_col = 1
            
            for source_col in range(1, units_ws.max_column + 1):
                header = units_ws.cell(row=1, column=source_col).value
                if header and str(header).strip():
                    header_str = str(header).strip()
                    col_mapping[source_col] = target_col
                    hier_ws.cell(row=headers_row, column=target_col, value=header_str)
                    target_col += 1
            
            # Now create dynamic formulas that reference Units sheet
            legacy_row = headers_row + 1
            
            # Create formulas for each row in Units sheet
            for units_data_row in range(2, units_row):
                for source_col, target_col in col_mapping.items():
                    # Create formula that references Units sheet
                    formula = f"=Units.{self.get_column_letter(source_col)}{units_data_row}"
                    hier_ws.cell(row=legacy_row, column=target_col, value=formula)
                
                legacy_row += 1
            
            # Apply formatting to make it look like traditional view
            # Bold the headers
            for col in range(1, target_col):
                cell = hier_ws.cell(row=headers_row, column=col)
                cell.font = openpyxl.styles.Font(bold=True)
                
            # Auto-width columns
            for column in hier_ws.columns:
                max_length = 0
                column_letter = self.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                hier_ws.column_dimensions[column_letter].width = adjusted_width
            
            self.log_message("✅ Legacy View sheet created with complete dynamic column structure")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating dynamic Legacy View sheet: {str(e)}")
            return False
    
    def get_column_letter(self, col_num):
        """Convert column number to Excel letter (A, B, C, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result
            
            # Remove any existing legacy view sheets
            sheets_to_remove = ["Hierarchical View", "Hierarchical View1", "Legacy View"]
            for sheet_name in sheets_to_remove:
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]
                
            hier_ws = workbook.create_sheet("Legacy View")
            
            # Add title and instructions
            hier_ws.cell(row=1, column=1, value="Legacy View - Dynamic Building Structure")
            hier_ws.cell(row=2, column=1, value="(Auto-updates from Units sheet - formulas link to live data)")
            
            # Set up headers in row 4
            headers_row = 4
            
            # Create column mapping and copy headers
            col_mapping = {}
            target_col = 1
            
            for source_col in range(1, units_ws.max_column + 1):
                header = units_ws.cell(row=1, column=source_col).value
                if header and str(header).strip():
                    header_str = str(header).strip()
                    col_mapping[source_col] = target_col
                    hier_ws.cell(row=headers_row, column=target_col, value=header_str)
                    target_col += 1
            
            # Create formulas that reference Units sheet data
            legacy_row = headers_row + 1
            
            for units_data_row in range(2, units_row):
                for source_col, target_col in col_mapping.items():
                    # Create formula that dynamically references Units sheet
                    col_letter = self.get_column_letter(source_col)
                    formula = f"=Units.{col_letter}{units_data_row}"
                    hier_ws.cell(row=legacy_row, column=target_col, value=formula)
                
                legacy_row += 1
            
            # Format headers
            for col in range(1, len(col_mapping) + 1):
                cell = hier_ws.cell(row=headers_row, column=col)
                if hasattr(openpyxl.styles, 'Font'):
                    cell.font = openpyxl.styles.Font(bold=True)
                
            # Auto-width columns
            try:
                for column in hier_ws.columns:
                    max_length = 0
                    column_letter = self.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    hier_ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as e:
                self.log_message(f"⚠️ Column auto-width failed: {e}")
            
            self.log_message("✅ Legacy View sheet created with dynamic formulas")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating dynamic Legacy View sheet: {str(e)}")
            return False
                                hier_ws.cell(row=unit_row, column=target_col, value=source_value)
                        
                        current_row += 1
                
                # Add small space between buildings
                current_row += 1
            
            # Add note about the view
            note_row = current_row + 2
            hier_ws.cell(row=note_row, column=1, 
                        value="Legacy View - Traditional building/unit structure")
            hier_ws.cell(row=note_row + 1, column=1, 
                        value="Shows all original data columns in hierarchical format")
            
            # Style the title
            title_cell = hier_ws.cell(row=1, column=1)
            title_cell.font = Font(size=16, bold=True)
            
            # Style headers - get the actual number of columns we have
            max_col = target_col - 1  # target_col is now one past the last column we used
            for col in range(1, max_col + 1):
                header_cell = hier_ws.cell(row=headers_row, column=col)
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Auto-width columns
            for column in hier_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                hier_ws.column_dimensions[column_letter].width = adjusted_width
            
            self.log_message("✅ Legacy View sheet created with complete column structure")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating Legacy View sheet: {str(e)}")
            return False

    def create_hierarchical_view_sheet_from_files(self, input_path, output_path, building_names):
        """Create a Legacy View sheet by opening the output file and adding the sheet"""
        try:
            self.log_message("📋 Creating Legacy View sheet from files...")
            
            # Open the output file (that should already have Units sheet)
            workbook = openpyxl.load_workbook(output_path)
            
            # Get the Units sheet to understand its structure
            units_ws = workbook['Units']
            units_row = units_ws.max_row + 1  # +1 to account for next available row
            
            # Call the main legacy view creation method
            success = self.create_hierarchical_view_sheet(workbook, units_ws, units_row)
            
            if success:
                # Save the updated workbook
                workbook.save(output_path)
                workbook.close()
                self.log_message("✅ Legacy View sheet added successfully")
                return True
            else:
                workbook.close()
                return False
                
        except Exception as e:
            self.log_message(f"❌ Error creating Legacy View sheet from files: {str(e)}")
            return False

    def find_column_letter(self, worksheet, header_name):
        """Find the column letter for a given header name in the worksheet"""
        try:
            # Check first row for headers
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == header_name:
                    return get_column_letter(col)
            return None
        except:
            return None

    def analyze_bank_schedule_data(self, file_path):
        """Analyze the Bank Schedule data and return counts of buildings, units, empty rows, and building names."""
        try:
            # Read the Bank Schedule sheet, starting from row 3 (index 2)
            df = pd.read_excel(file_path, sheet_name="Bank Schedule", header=2)
            
            # Get column G (Unit Type) - this should be index 6 if we count from A=0
            # But since headers start at row 3, we need to check the actual column name
            unit_type_col = None
            for col in df.columns:
                if 'Unit Type' in str(col) or col == 'G' or df.columns.get_loc(col) == 6:
                    unit_type_col = col
                    break
            
            if unit_type_col is None:
                # If we can't find by name, try by position (column G = index 6)
                if len(df.columns) > 6:
                    unit_type_col = df.columns[6]
                else:
                    raise ValueError("Could not find 'Unit Type' column (column G)")
            
            self.log_message(f"Using column '{unit_type_col}' as Unit Type column")
            
            building_count = 0
            unit_count = 0
            empty_count = 0
            building_names = []
            current_building = None
            
            # Iterate through all rows
            for index, row in df.iterrows():
                unit_type_value = row[unit_type_col]
                
                # Check if row is empty (Unit Type column is empty/null)
                if pd.isna(unit_type_value) or str(unit_type_value).strip() == '':
                    empty_count += 1
                    continue
                
                # Check if this is a building row
                # Building rows have value in Unit Type but should have no other significant data
                # Let's check if most other columns are empty for this row
                non_empty_cols = 0
                for col in df.columns:
                    if col != unit_type_col and not pd.isna(row[col]) and str(row[col]).strip() != '':
                        non_empty_cols += 1
                
                # If this row has Unit Type but few other fields, it's likely a building
                if non_empty_cols <= 2:  # Allow for some tolerance
                    building_name = str(unit_type_value).strip()
                    # Filter out notes and other non-building entries
                    if not any(keyword in building_name.lower() for keyword in ['note', 'capital values', 'rent pa', 'as a result']):
                        building_count += 1
                        building_names.append(building_name)
                        current_building = building_name
                        self.log_message(f"Found building: {current_building}")
                else:
                    # This is a unit row (has Unit Type and other data)
                    unit_count += 1
                    if current_building:
                        self.log_message(f"  Unit in {current_building}: {unit_type_value}")
            
            self.log_message(f"Analysis complete: {building_count} buildings, {unit_count} units")
            
            return {
                'buildings': building_count,
                'units': unit_count,
                'empty_rows': empty_count,
                'total_rows': len(df),
                'building_names': building_names
            }
            
        except Exception as e:
            self.log_message(f"Error analyzing data: {str(e)}")
            raise
            
    def create_buildings_summary_sheet(self, input_path, output_path, analysis_results):
        """Create a new 'Buildings' sheet with building summary data while preserving original formatting."""
        try:
            self.log_message("Step 4: Creating Buildings summary sheet...")
            
            # Import openpyxl for direct Excel manipulation
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            import shutil
            
            # First, copy the original file to preserve all formatting and formulas
            shutil.copy2(input_path, output_path)
            self.log_message("Original file copied with all formatting preserved")
            
            # Now open the copied file and add the new Buildings sheet
            workbook = load_workbook(output_path)
            
            # Get building names from analysis results
            building_names = analysis_results['building_names']
            
            self.log_message(f"Creating Buildings summary for {len(building_names)} buildings")
            
            # Create the new Buildings worksheet
            buildings_ws = workbook.create_sheet("Buildings")
            
            # Define headers
            headers = [
                'Building', 'Net Area', 'Rent PA (£)', '2023 ERV (£)', 
                '2024 ERV (£)', 'ERV 2024 £.Sq.ft', 'ERV Variation', '2024 Cap Valn. (£)'
            ]
            
            # Write headers to the first row
            for col_num, header in enumerate(headers, 1):
                buildings_ws.cell(row=1, column=col_num, value=header)
            
            # Write building names to the first column (starting from row 2)
            for row_num, building_name in enumerate(building_names, 2):
                buildings_ws.cell(row=row_num, column=1, value=building_name)
            
            # Add autofilter to the headers
            buildings_ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
            self.log_message("✅ Autofilter applied to Buildings sheet headers")
            
            # Auto-width all columns
            for column in buildings_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                buildings_ws.column_dimensions[column_letter].width = adjusted_width
            self.log_message("✅ Auto-width applied to all columns in Buildings sheet")
            
            # Save the workbook with the new sheet
            workbook.save(output_path)
            workbook.close()
            
            self.log_message("✅ Buildings summary sheet created successfully")
            self.log_message("✅ Original formatting and formulas preserved in all existing sheets")
            return True
            
        except Exception as e:
            self.log_message(f"❌ Error creating Buildings summary sheet: {str(e)}")
            raise

    def create_units_sheet(self, input_path, output_path, building_names):
        """Create a new 'Units' sheet with all unit data and building validation dropdown."""
        try:
            self.log_message("Step 5: Creating Units sheet with data validation...")
            
            # Import additional openpyxl components
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import NamedStyle
            from openpyxl.worksheet.datavalidation import DataValidation
            import copy
            
            # Open the workbook
            workbook = load_workbook(output_path)
            
            # Read the original Bank Schedule sheet with openpyxl to preserve formatting
            source_ws = workbook["Bank Schedule"]
            
            # Also read with pandas for data analysis
            df = pd.read_excel(input_path, sheet_name="Bank Schedule", header=2)
            
            # Find the Unit Type column
            unit_type_col = None
            unit_type_col_idx = None
            for col in df.columns:
                if 'Unit Type' in str(col) or col == 'G' or df.columns.get_loc(col) == 6:
                    unit_type_col = col
                    unit_type_col_idx = df.columns.get_loc(col)
                    break
            
            if unit_type_col is None and len(df.columns) > 6:
                unit_type_col = df.columns[6]
                unit_type_col_idx = 6
            
            # Create the new Units worksheet
            units_ws = workbook.create_sheet("Units")
            
            # Copy headers from source sheet (row 3, which is index 2)
            # First add "Building" as the first column
            units_ws.cell(row=1, column=1, value="Building")
            
            # Then copy all original headers (shifting by 1 column)
            header_row = 3  # Headers are in row 3 of source sheet
            for col_idx in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=header_row, column=col_idx)
                target_cell = units_ws.cell(row=1, column=col_idx + 1)
                
                # Handle potential formulas in headers (rare but possible)
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    refactored_formula = self.refactor_formula(
                        source_cell.value, 
                        header_row, 
                        1, 
                        column_offset=1
                    )
                    target_cell.value = refactored_formula
                else:
                    target_cell.value = source_cell.value
                    
                # Copy formatting safely
                try:
                    if source_cell.has_style:
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy.copy(source_cell.protection)
                        target_cell.alignment = copy.copy(source_cell.alignment)
                except Exception as e:
                    self.log_message(f"   ⚠️ Could not copy formatting for header cell: {str(e)}")
            
            # Track units and their buildings
            current_building = None
            units_row = 2  # Start from row 2 (after headers)
            
            # Process all data rows starting from row 4 (after headers)
            for source_row_idx in range(4, source_ws.max_row + 1):
                # Get the unit type value
                unit_type_cell = source_ws.cell(row=source_row_idx, column=unit_type_col_idx + 1)  # +1 because Excel is 1-indexed
                unit_type_value = unit_type_cell.value
                
                # Skip empty rows
                if not unit_type_value or str(unit_type_value).strip() == '':
                    continue
                
                # Check if this is a building row or unit row
                non_empty_cols = 0
                for col_idx in range(1, source_ws.max_column + 1):
                    if col_idx != unit_type_col_idx + 1:  # Skip the unit type column
                        cell_value = source_ws.cell(row=source_row_idx, column=col_idx).value
                        if cell_value and str(cell_value).strip() != '':
                            non_empty_cols += 1
                
                # If this row has Unit Type but few other fields, it's likely a building
                if non_empty_cols <= 2:
                    building_name = str(unit_type_value).strip()
                    # Filter out notes and other non-building entries
                    if not any(keyword in building_name.lower() for keyword in ['note', 'capital values', 'rent pa', 'as a result']):
                        current_building = building_name
                else:
                    # This is a unit row - copy all data
                    if current_building:
                        # Set building name in first column
                        units_ws.cell(row=units_row, column=1, value=current_building)
                        
                        # Copy all other data (shifting by 1 column)
                        for col_idx in range(1, source_ws.max_column + 1):
                            source_cell = source_ws.cell(row=source_row_idx, column=col_idx)
                            target_cell = units_ws.cell(row=units_row, column=col_idx + 1)
                            
                            # Handle formulas with refactoring
                            if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                # Refactor formula for new sheet structure
                                refactored_formula = self.refactor_formula(
                                    source_cell.value, 
                                    source_row_idx, 
                                    units_row, 
                                    column_offset=1
                                )
                                target_cell.value = refactored_formula
                            else:
                                # Copy regular values
                                target_cell.value = source_cell.value
                            
                            # Copy formatting safely
                            try:
                                if source_cell.has_style:
                                    target_cell.font = copy.copy(source_cell.font)
                                    target_cell.border = copy.copy(source_cell.border)
                                    target_cell.fill = copy.copy(source_cell.fill)
                                    target_cell.number_format = source_cell.number_format
                                    target_cell.protection = copy.copy(source_cell.protection)
                                    target_cell.alignment = copy.copy(source_cell.alignment)
                            except Exception as e:
                                pass  # Continue if formatting copy fails
                        
                        units_row += 1
            
            self.log_message(f"✅ Copied {units_row - 2} unit rows to Units sheet")
            
            # Add Status column at the end
            self.log_message("📝 Adding Status column with data validation...")
            
            # Find the "Start Date" column in the original sheet
            start_date_col_idx = None
            df = pd.read_excel(input_path, sheet_name="Bank Schedule", header=2)
            
            for col_idx, col_name in enumerate(df.columns):
                # Clean up column name thoroughly: remove extra spaces, convert to lowercase
                clean_col_name = ' '.join(str(col_name).split()).lower() if col_name else ''
                if 'start date' == clean_col_name or 'start date' in clean_col_name:
                    start_date_col_idx = col_idx
                    break
            
            if start_date_col_idx is None:
                # Fallback to column M (index 12) if not found by header
                start_date_col_idx = 12
                self.log_message(f"   ⚠️ 'Start Date' column not found by header, using column M (index {start_date_col_idx})")
            else:
                self.log_message(f"   ✅ Found 'Start Date' column at index {start_date_col_idx} ('{df.columns[start_date_col_idx]}')")
            
            # Add Status header
            status_col = units_ws.max_column + 1
            status_header_cell = units_ws.cell(row=1, column=status_col, value="Status")
            
            # Copy header formatting from adjacent cell
            try:
                adjacent_cell = units_ws.cell(row=1, column=status_col - 1)
                if adjacent_cell.has_style:
                    status_header_cell.font = copy.copy(adjacent_cell.font)
                    status_header_cell.border = copy.copy(adjacent_cell.border)
                    status_header_cell.fill = copy.copy(adjacent_cell.fill)
                    status_header_cell.alignment = copy.copy(adjacent_cell.alignment)
            except Exception as e:
                pass
            
            # Populate Status column based on Start Date
            for row in range(2, units_row):  # Skip header row
                # Get the Start Date value from the corresponding column (adjusted for Building column shift)
                start_date_cell = units_ws.cell(row=row, column=start_date_col_idx + 2)  # +2 because: +1 for Building column, +1 for 0-based to 1-based
                start_date_value = start_date_cell.value
                
                # Create Status cell
                status_cell = units_ws.cell(row=row, column=status_col)
                
                # Set Status based on Start Date
                if start_date_value is None or str(start_date_value).strip() == '' or str(start_date_value).lower() == 'nan':
                    status_cell.value = "Void"
                else:
                    status_cell.value = "Let"
            
            self.log_message(f"✅ Status column populated for {units_row - 2} units")
            
            # Create Excel-compatible data validation for Status column
            try:
                from openpyxl.worksheet.datavalidation import DataValidation
                
                # Use Excel's exact format from reverse engineering
                status_formula = '"Let, Void, Under Ref, Mothballed, Under Off"'
                
                # Create data validation matching Excel's structure
                dv = DataValidation(
                    type="list",
                    formula1=status_formula,
                    allowBlank=True,  # Excel uses allowBlank="1"
                    showInputMessage=True,  # Excel uses showInputMessage="1"
                    showErrorMessage=True   # Excel uses showErrorMessage="1"
                )
                # No custom error/prompt messages (Excel's version is clean)
                
                # Apply to the entire Status column (like Excel's A2:A1048576)
                if units_row > 2:  # Only if we have data rows
                    status_col_letter = get_column_letter(status_col)
                    range_ref = f"{status_col_letter}2:{status_col_letter}1048576"
                    dv.add(range_ref)
                    units_ws.add_data_validation(dv)
                    self.log_message(f"✅ Excel-compatible data validation applied to Status column ({status_col_letter}2:{status_col_letter}{units_row - 1})")
                
            except Exception as e:
                self.log_message(f"   ⚠️ Could not add data validation to Status column: {str(e)}")
                self.log_message("   Status column created without dropdown validation")
            
            # Skip data validation dropdown to avoid incomplete building lists
            self.log_message(f"📝 Building column populated for all {len(building_names)} buildings (no dropdown to avoid incomplete lists)")
            
            # Add autofilter to the headers (including new Status column)
            units_ws.auto_filter.ref = f"A1:{get_column_letter(units_ws.max_column)}{units_row - 1}"
            self.log_message("✅ Autofilter applied to Units sheet headers (including Status column)")
            
            # Auto-width all columns (with better error handling)
            try:
                for column in units_ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            cell_value = cell.value
                            if cell_value is not None:
                                cell_length = len(str(cell_value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    # Set reasonable width limits
                    adjusted_width = max(min(max_length + 2, 50), 8)  # Min 8, Max 50
                    units_ws.column_dimensions[column_letter].width = adjusted_width
                self.log_message("✅ Auto-width applied to all columns in Units sheet")
            except Exception as e:
                self.log_message(f"⚠️ Could not apply auto-width: {str(e)}")
            
            # Save the workbook with better error handling
            try:
                workbook.save(output_path)
                workbook.close()
                self.log_message("✅ Units sheet created successfully (no dropdown for complete building list)")
                return True
            except Exception as e:
                self.log_message(f"❌ Error saving workbook: {str(e)}")
                workbook.close()
                raise
            
        except Exception as e:
            self.log_message(f"❌ Error creating Units sheet: {str(e)}")
            raise

    def sanitize_file(self):
        """Process the selected Excel file and save the sanitized version."""
        try:
            self.log_message("\n" + "="*50)
            self.log_message("Starting sanitization process...")
            
            input_path = self.input_file_path.get()
            
            # Validate input file
            is_valid, validation_message = self.validate_excel_file(input_path)
            if not is_valid:
                messagebox.showerror("Error", validation_message)
                return
                
            self.log_message("Step 1: Input file validation completed.")
            
            # Analyze the Bank Schedule data
            self.log_message("Step 2: Analyzing Bank Schedule data structure...")
            
            try:
                analysis_results = self.analyze_bank_schedule_data(input_path)
                
                self.log_message("\n" + "-"*30)
                self.log_message("📊 DATA ANALYSIS RESULTS:")
                self.log_message("-"*30)
                self.log_message(f"🏢 Buildings found: {analysis_results['buildings']}")
                self.log_message(f"🏠 Units found: {analysis_results['units']}")
                self.log_message(f"📝 Empty rows: {analysis_results['empty_rows']}")
                self.log_message(f"📋 Total rows processed: {analysis_results['total_rows']}")
                self.log_message("-"*30 + "\n")
                
            except Exception as e:
                self.log_message(f"❌ Error during analysis: {str(e)}")
                messagebox.showerror("Analysis Error", f"Could not analyze data structure: {str(e)}")
                return
            
            self.log_message("Step 3: Processing file and creating Buildings summary...")
            
            # Show save dialog
            input_name = os.path.basename(input_path)
            suggested_name = f"sanitized_{input_name}"
            
            output_path = filedialog.asksaveasfilename(
                title="Save Sanitized File As",
                defaultextension=".xlsx",
                initialfile=suggested_name,
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*")
                ]
            )
            
            if not output_path:
                self.log_message("⚠ Save operation cancelled by user.")
                self.log_message("="*50 + "\n")
                return
            
            # Create the new file with Buildings summary sheet
            try:
                self.create_buildings_summary_sheet(input_path, output_path, analysis_results)
                # Create Units sheet using the building names from analysis
                self.create_units_sheet(input_path, output_path, analysis_results['building_names'])
                # Create hierarchical legacy view for familiar format
                self.create_hierarchical_view_sheet_from_files(input_path, output_path, analysis_results['building_names'])
            except Exception as e:
                self.log_message(f"❌ Error creating output file: {str(e)}")
                messagebox.showerror("File Creation Error", f"Could not create output file: {str(e)}")
                return
            
            self.log_message(f"✅ Sanitized file saved successfully to: {os.path.basename(output_path)}")
            
            # Show success message with analysis results
            success_message = f"""File processed successfully!

Analysis Results:
• Buildings: {analysis_results['buildings']}
• Units: {analysis_results['units']}
• Empty rows: {analysis_results['empty_rows']}
• Total rows: {analysis_results['total_rows']}

New Features:
• Added 'Buildings' summary sheet with autofilter and auto-width
• Added 'Units' sheet with all unit data and building validation dropdown
• Extracted building names automatically
• Preserved all original formatting and formulas

Saved to:
{output_path}"""
            
            messagebox.showinfo("Success", success_message)
            
            self.log_message("="*50 + "\n")
            
        except Exception as e:
            error_msg = f"Error during sanitization: {str(e)}"
            self.log_message(f"✗ {error_msg}")
            messagebox.showerror("Error", error_msg)


def main():
    """Main function to run the application."""
    root = tk.Tk()
    app = BankScheduleSanitizer(root)
    root.mainloop()


if __name__ == "__main__":
    main()
