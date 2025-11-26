#!/usr/bin/env python3
import openpyxl

def search_for_missing_building():
    """Search for 50 Eastcastle Street in both Bank Schedule and Units sheets"""
    
    print("🔍 Searching for '50 Eastcastle Street'...")
    
    # Check Bank Schedule
    print("\n📋 Checking Bank Schedule sheet:")
    try:
        bank_wb = openpyxl.load_workbook('data/Leasing Bank Schedule June 2025.xlsx', data_only=False)
        bank_ws = bank_wb['Bank Schedule']
        
        found_in_bank = []
        for row in range(1, 500):
            property_val = bank_ws.cell(row=row, column=5).value  # Column E (Property)
            if property_val and "50 eastcastle" in str(property_val).lower():
                unit_demise = bank_ws.cell(row=row, column=6).value  # Column F (Unit Demise)
                cap_valn = bank_ws.cell(row=row, column=28).value  # Column AB (Cap Valn)
                found_in_bank.append({
                    'row': row,
                    'property': property_val,
                    'unit_demise': unit_demise,
                    'cap_valn': cap_valn
                })
        
        if found_in_bank:
            print(f"  ✅ Found {len(found_in_bank)} rows with '50 Eastcastle':")
            for match in found_in_bank:
                print(f"    Row {match['row']}: '{match['property']}' | Unit: '{match['unit_demise']}' | Cap Valn: '{match['cap_valn']}'")
        else:
            print("  ❌ No '50 Eastcastle Street' found in Bank Schedule")
        
        bank_wb.close()
        
    except Exception as e:
        print(f"  ❌ Error checking Bank Schedule: {e}")
    
    # Check Units sheet
    print("\n📋 Checking Units sheet:")
    try:
        units_wb = openpyxl.load_workbook('data/27.xlsx', data_only=False)
        units_ws = units_wb['Units']
        
        found_in_units = []
        for row in range(1, 200):
            for col in range(1, 20):
                cell_val = units_ws.cell(row=row, column=col).value
                if cell_val and "50 eastcastle" in str(cell_val).lower():
                    found_in_units.append({
                        'row': row,
                        'col': col,
                        'value': cell_val
                    })
        
        if found_in_units:
            print(f"  ✅ Found {len(found_in_units)} references to '50 Eastcastle' in Units:")
            for match in found_in_units:
                print(f"    Row {match['row']}, Col {match['col']}: '{match['value']}'")
        else:
            print("  ❌ No '50 Eastcastle Street' found in Units sheet")
        
        units_wb.close()
        
    except Exception as e:
        print(f"  ❌ Error checking Units sheet: {e}")
    
    # Also search for similar patterns
    print("\n🔍 Searching for buildings with '0' values in Buildings sheet:")
    try:
        buildings_wb = openpyxl.load_workbook('data/27.xlsx', data_only=True)  # Use data_only to see calculated values
        buildings_ws = buildings_wb['Buildings']
        
        zero_value_buildings = []
        for row in range(2, 50):
            building_name = buildings_ws.cell(row=row, column=1).value
            if building_name:
                # Check Net Area column (column 2)
                net_area = buildings_ws.cell(row=row, column=2).value
                # Check Rent PA column (column 3) 
                rent_pa = buildings_ws.cell(row=row, column=3).value
                # Check Cap Valn column (column 6)
                cap_valn = buildings_ws.cell(row=row, column=6).value
                
                if (net_area == 0 or rent_pa == 0) and not cap_valn:
                    zero_value_buildings.append({
                        'row': row,
                        'name': building_name,
                        'net_area': net_area,
                        'rent_pa': rent_pa,
                        'cap_valn': cap_valn
                    })
        
        if zero_value_buildings:
            print(f"  ⚠️ Found {len(zero_value_buildings)} buildings with 0 values:")
            for building in zero_value_buildings:
                print(f"    Row {building['row']}: {building['name']}")
                print(f"      Net Area: {building['net_area']}, Rent PA: {building['rent_pa']}, Cap Valn: {building['cap_valn']}")
        else:
            print("  ✅ No buildings with 0 values found")
        
        buildings_wb.close()
        
    except Exception as e:
        print(f"  ❌ Error checking Buildings sheet: {e}")

if __name__ == "__main__":
    search_for_missing_building()
