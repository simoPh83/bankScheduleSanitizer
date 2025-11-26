#!/usr/bin/env python3

import openpyxl
import os

def get_column_letter(col_num):
    """Convert column number to Excel letter (A, B, C, etc.)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result

def fix_legacy_view_formulas(file_path):
    """Fix the Legacy View formulas to use correct Excel format"""
    try:
        print(f"Opening file: {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        
        # Check if Legacy View sheet exists
        if "Legacy View" not in workbook.sheetnames:
            print("Error: 'Legacy View' sheet not found")
            return False
            
        if "Units" not in workbook.sheetnames:
            print("Error: 'Units' sheet not found")
            return False
            
        legacy_ws = workbook["Legacy View"]
        units_ws = workbook["Units"]
        
        print("Fixing Legacy View formulas...")
        
        # Find the headers row (should be row 4)
        headers_row = 4
        data_start_row = headers_row + 1
        
        # Create column mapping
        col_mapping = {}
        target_col = 1
        
        for source_col in range(1, units_ws.max_column + 1):
            header = units_ws.cell(row=1, column=source_col).value
            if header and str(header).strip():
                col_mapping[source_col] = target_col
                target_col += 1
        
        # Fix formulas for all data rows
        max_row = legacy_ws.max_row
        fixed_count = 0
        
        for legacy_row in range(data_start_row, max_row + 1):
            # Calculate corresponding Units row
            units_data_row = legacy_row - data_start_row + 2  # +2 because Units starts at row 2
            
            for source_col, target_col in col_mapping.items():
                cell = legacy_ws.cell(row=legacy_row, column=target_col)
                if cell.value and str(cell.value).startswith("="):
                    # Create correct formula format
                    col_letter = get_column_letter(source_col)
                    correct_formula = f"=Units!{col_letter}{units_data_row}"
                    
                    # Only update if it's different
                    if cell.value != correct_formula:
                        cell.value = correct_formula
                        fixed_count += 1
        
        print(f"Fixed {fixed_count} formulas")
        
        # Save the corrected file
        backup_path = file_path.replace('.xlsx', '_backup_before_formula_fix.xlsx')
        print(f"Creating backup: {backup_path}")
        workbook.save(backup_path)
        
        # Save the corrected version
        workbook.save(file_path)
        workbook.close()
        
        print("✅ Legacy View formulas fixed successfully!")
        print(f"✅ File saved: {file_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error fixing formulas: {str(e)}")
        return False

if __name__ == "__main__":
    # Fix the most recent file
    file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/8.xlsx"
    
    if os.path.exists(file_path):
        fix_legacy_view_formulas(file_path)
    else:
        print(f"File not found: {file_path}")
