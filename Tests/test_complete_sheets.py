#!/usr/bin/env python3
"""
Test script to verify the Buildings and Units sheet creation functionality
"""

import pandas as pd
import sys
import os
import tempfile

# Add the current directory to path so we can import our main module
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from bank_schedule_sanitizer import BankScheduleSanitizer
import tkinter as tk

def test_complete_sheet_creation():
    """Test the complete Buildings and Units sheet creation functionality."""
    
    print("Testing Complete Sheet Creation (Buildings + Units)...")
    print("="*60)
    
    test_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/Leasing Bank Schedule June 2025.xlsx"
    
    if not os.path.exists(test_file):
        print(f"❌ Test file not found: {test_file}")
        return False
    
    try:
        # Create a minimal tkinter instance (needed for the class)
        root = tk.Tk()
        root.withdraw()  # Hide the window for testing
        
        # Create an instance of our sanitizer
        sanitizer = BankScheduleSanitizer(root)
        
        # Override the log_message method to print to console for testing
        def test_log(message):
            print(f"[LOG] {message}")
        sanitizer.log_message = test_log
        
        print("📊 Running analysis...")
        
        # Test the analysis function first
        analysis_results = sanitizer.analyze_bank_schedule_data(test_file)
        
        print(f"\n🎉 Analysis completed:")
        print(f"• Buildings: {analysis_results['buildings']}")
        print(f"• Units: {analysis_results['units']}")
        print(f"• Empty rows: {analysis_results['empty_rows']}")
        print(f"• Total rows: {analysis_results['total_rows']}")
        
        # Create a temporary output file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_output = tmp.name
        
        print(f"\n🏗️  Testing complete sheet creation...")
        print(f"Output file: {temp_output}")
        
        # Test the complete sheet creation process
        buildings_result = sanitizer.create_buildings_summary_sheet(test_file, temp_output, analysis_results)
        units_result = sanitizer.create_units_sheet(test_file, temp_output, sanitizer.building_names)
        
        if buildings_result and units_result:
            print("✅ Both Buildings and Units sheets created successfully")
            
            # Verify the output file
            print("\n🔍 Verifying output file...")
            
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(temp_output)
                sheet_names = workbook.sheet_names
                
                print(f"📋 Sheets in output file: {sheet_names}")
                
                # Check Buildings sheet
                if 'Buildings' in sheet_names:
                    print("✅ 'Buildings' sheet found")
                    buildings_ws = workbook['Buildings']
                    print(f"   📊 Buildings sheet dimensions: {buildings_ws.max_row} rows x {buildings_ws.max_column} columns")
                    print(f"   🎛️  Autofilter applied: {buildings_ws.auto_filter.ref is not None}")
                else:
                    print("❌ 'Buildings' sheet missing")
                
                # Check Units sheet
                if 'Units' in sheet_names:
                    print("✅ 'Units' sheet found")
                    units_ws = workbook['Units']
                    print(f"   📊 Units sheet dimensions: {units_ws.max_row} rows x {units_ws.max_column} columns")
                    print(f"   🎛️  Autofilter applied: {units_ws.auto_filter.ref is not None}")
                    
                    # Check data validation
                    data_validations = list(units_ws.data_validations)
                    if data_validations:
                        dv = data_validations[0]
                        print(f"   ✅ Data validation found: {dv.sqref}")
                        print(f"   📝 Validation formula: {dv.formula1[:100]}...")
                    else:
                        print("   ⚠️  No data validation found")
                    
                    # Check first few rows
                    print("   📋 Sample data:")
                    for row in range(1, min(4, units_ws.max_row + 1)):
                        building = units_ws.cell(row=row, column=1).value
                        unit_type = units_ws.cell(row=row, column=8).value  # Assuming Unit Type is around column 8
                        print(f"      Row {row}: Building='{building}', Unit Type='{unit_type}'")
                else:
                    print("❌ 'Units' sheet missing")
                
                workbook.close()
                
                # Check original sheet preservation
                original_sheets = ['Bank Schedule', 'Voids Tab', 'Tenant Trade', 'Lease Events ', 'Calculations ']
                preserved = [sheet for sheet in original_sheets if sheet in sheet_names]
                print(f"📄 Original sheets preserved: {len(preserved)}/{len(original_sheets)}")
                if len(preserved) == len(original_sheets):
                    print("✅ All original sheets preserved")
                else:
                    print("⚠️  Some original sheets missing")
                    
            except Exception as e:
                print(f"❌ Error verifying file: {e}")
                return False
        else:
            print("❌ Sheet creation failed")
            return False
        
        # Clean up
        if os.path.exists(temp_output):
            os.unlink(temp_output)
            print("🧹 Temporary file cleaned up")
        
        root.destroy()
        
        print(f"\n🎉 All tests passed! Complete sheet creation is working correctly.")
        return True
        
    except Exception as e:
        print(f"❌ Error during testing: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_complete_sheet_creation()
    sys.exit(0 if success else 1)
