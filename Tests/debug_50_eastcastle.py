#!/usr/bin/env python3
import openpyxl

def debug_specific_building():
    """Debug why 50 Eastcastle Street has no Cap Valn formula"""
    
    print("🔍 Debugging '50 Eastcastle Street'...")
    
    # Check Bank Schedule for this specific building
    try:
        bank_wb = openpyxl.load_workbook('data/Leasing Bank Schedule June 2025.xlsx', data_only=False)
        bank_ws = bank_wb['Bank Schedule']
        
        print("\n📋 Searching Bank Schedule for '50 Eastcastle Street':")
        found_matches = []
        
        # Search through Property column (E) and other relevant columns
        for row in range(1, bank_ws.max_row + 1):
            property_val = bank_ws.cell(row=row, column=5).value  # Column E (Property)
            unit_demise = bank_ws.cell(row=row, column=6).value   # Column F (Unit Demise)
            
            # Check for exact match or partial match
            if property_val and "50 eastcastle" in str(property_val).lower():
                cap_valn = bank_ws.cell(row=row, column=28).value  # Column AB (Cap Valn)
                found_matches.append({
                    'row': row,
                    'property': property_val,
                    'unit_demise': unit_demise,
                    'cap_valn': cap_valn
                })
        
        if found_matches:
            print(f"  ✅ Found {len(found_matches)} rows with '50 Eastcastle':")
            for match in found_matches:
                print(f"    Row {match['row']}: Property='{match['property']}'")
                print(f"                     Unit Demise='{match['unit_demise']}'")
                print(f"                     Cap Valn='{match['cap_valn']}'")
                print()
        else:
            print("  ❌ No exact matches found")
            
            # Try broader search
            print("  🔍 Trying broader search for 'eastcastle' + '50':")
            for row in range(1, bank_ws.max_row + 1):
                property_val = bank_ws.cell(row=row, column=5).value
                if property_val and "eastcastle" in str(property_val).lower() and "50" in str(property_val):
                    unit_demise = bank_ws.cell(row=row, column=6).value
                    cap_valn = bank_ws.cell(row=row, column=28).value
                    print(f"    Row {row}: Property='{property_val}'")
                    print(f"                     Unit Demise='{unit_demise}'")
                    print(f"                     Cap Valn='{cap_valn}'")
                    print()
        
        bank_wb.close()
        
    except Exception as e:
        print(f"  ❌ Error checking Bank Schedule: {e}")
        import traceback
        traceback.print_exc()
    
    # Check what our Cap Valn mapping found
    print("\n📋 Checking what our Cap Valn analysis found:")
    print("(This will help us understand why '50 Eastcastle Street' was missed)")
    
    # Check if this building appears in the Units sheet and how
    try:
        units_wb = openpyxl.load_workbook('data/27.xlsx', data_only=False)
        units_ws = units_wb['Units']
        
        print("\n📋 Checking Units sheet for '50 Eastcastle Street':")
        
        # Find Property column in Units sheet
        property_col = None
        for col in range(1, 20):
            header_val = units_ws.cell(row=1, column=col).value
            if header_val and str(header_val).lower().strip() == "property":
                property_col = col
                break
        
        if property_col:
            print(f"  Property column found at: {property_col}")
            found_units = []
            
            for row in range(2, units_ws.max_row + 1):
                property_val = units_ws.cell(row=row, column=property_col).value
                if property_val and "50 eastcastle" in str(property_val).lower():
                    found_units.append({
                        'row': row,
                        'property': property_val
                    })
            
            if found_units:
                print(f"  ✅ Found {len(found_units)} units for '50 Eastcastle Street':")
                for unit in found_units:
                    print(f"    Row {unit['row']}: '{unit['property']}'")
            else:
                print("  ❌ No units found for '50 Eastcastle Street'")
        else:
            print("  ❌ Property column not found in Units sheet")
        
        units_wb.close()
        
    except Exception as e:
        print(f"  ❌ Error checking Units sheet: {e}")

if __name__ == "__main__":
    debug_specific_building()
