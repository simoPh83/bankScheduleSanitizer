#!/usr/bin/env python3
import openpyxl

def check_cap_valn_formulas():
    """Check if the Cap Valn formulas in Buildings sheet are working correctly"""
    
    try:
        # Load the latest generated file
        wb = openpyxl.load_workbook('data/27.xlsx', data_only=False)
        
        if 'Buildings' not in wb.sheetnames:
            print("❌ Buildings sheet not found")
            return
            
        buildings_ws = wb['Buildings']
        
        print("📋 Buildings Sheet - Cap Valn Column Check:")
        print("=" * 50)
        
        # Find the Cap Valn column
        cap_valn_col = None
        for col in range(1, 10):
            header_val = buildings_ws.cell(row=1, column=col).value
            if header_val and "2024 Cap Valn" in str(header_val):
                cap_valn_col = col
                print(f"🔍 Found Cap Valn column at: {chr(64+col)}")
                break
        
        if not cap_valn_col:
            print("❌ Cap Valn column not found")
            return
        
        # Check formulas in the Cap Valn column
        print(f"\n🔍 Cap Valn formulas and values (showing ALL buildings):")
        for row in range(2, 25):  # Check more buildings to see where they stop
            building_name = buildings_ws.cell(row=row, column=1).value
            if building_name:
                formula_cell = buildings_ws.cell(row=row, column=cap_valn_col)
                formula = formula_cell.value
                
                print(f"\nRow {row} - Building: {building_name}")
                print(f"  Formula: {formula}")
                
                # Also check with data_only to see calculated value
                wb_data = openpyxl.load_workbook('data/27.xlsx', data_only=True)
                buildings_data_ws = wb_data['Buildings']
                calculated_value = buildings_data_ws.cell(row=row, column=cap_valn_col).value
                print(f"  Calculated Value: {calculated_value}")
                wb_data.close()
            else:
                print(f"\nRow {row} - No building name found")
                break
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error checking Cap Valn formulas: {str(e)}")

if __name__ == "__main__":
    check_cap_valn_formulas()
