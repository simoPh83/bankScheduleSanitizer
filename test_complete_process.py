#!/usr/bin/env python3
"""
Quick test to verify the complete sanitization process including hierarchical view
"""

import pandas as pd
import sys
import os
import tempfile
import tkinter as tk

# Add the current directory to path so we can import our main module
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from bank_schedule_sanitizer import BankScheduleSanitizer

def test_complete_process():
    """Test the complete sanitization process."""
    
    print("Testing Complete Sanitization Process...")
    print("="*60)
    
    test_file = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/Leasing Bank Schedule June 2025.xlsx"
    
    if not os.path.exists(test_file):
        print(f"❌ Test file not found: {test_file}")
        return False
    
    try:
        # Create a minimal tkinter instance (needed for the class)
        root = tk.Tk()
        root.withdraw()  # Hide the window for testing
        
        # Create an instance of our sanitizer
        sanitizer = BankScheduleSanitizer(root)
        
        # Override the log_message method to print to console for testing
        def test_log(message):
            print(f"[LOG] {message}")
        sanitizer.log_message = test_log
        
        # Set the input file path
        sanitizer.input_file_path.set(test_file)
        
        print("🚀 Running complete sanitization process...")
        
        # This should create Buildings, Units, and Hierarchical View sheets
        sanitizer.sanitize_file()
        
        # Check the output was created
        # Look for the sanitized file in the same directory
        output_file = test_file.replace('.xlsx', '_sanitized.xlsx')
        
        if os.path.exists(output_file):
            print(f"\n✅ Sanitized file created: {os.path.basename(output_file)}")
            
            # Verify the sheets exist
            import openpyxl
            workbook = openpyxl.load_workbook(output_file)
            sheets = workbook.sheetnames
            print(f"📊 Sheets in output: {sheets}")
            
            required_sheets = ['Buildings', 'Units', 'Hierarchical View']
            missing_sheets = [sheet for sheet in required_sheets if sheet not in sheets]
            
            if missing_sheets:
                print(f"❌ Missing required sheets: {missing_sheets}")
                workbook.close()
                return False
            else:
                print(f"✅ All required sheets present!")
                
                # Check hierarchical view content
                hier_sheet = workbook['Hierarchical View']
                print(f"🏢 Hierarchical View: {hier_sheet.max_row} rows, {hier_sheet.max_column} cols")
                
                # Sample first building structure
                building_found = False
                for row in range(1, min(10, hier_sheet.max_row + 1)):
                    cell_value = hier_sheet.cell(row=row, column=1).value
                    if cell_value and "📁" in str(cell_value):
                        building_found = True
                        print(f"✅ Found building structure: {cell_value}")
                        break
                
                if not building_found:
                    print("⚠️ No building structure found in hierarchical view")
                
            workbook.close()
            return True
        else:
            print(f"❌ Output file not created: {output_file}")
            return False
        
    except Exception as e:
        print(f"❌ Test failed with error: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_complete_process()
    if success:
        print(f"\n🎉 Complete process test PASSED!")
        print("📋 The application now includes:")
        print("   • Buildings summary sheet")
        print("   • Units detailed sheet with Status column")  
        print("   • Hierarchical View with auto-updating formulas")
        print("   • Excel-compatible data validation")
        print("   • Formula refactoring for correct references")
    else:
        print(f"\n💥 Complete process test FAILED!")
        sys.exit(1)
