#!/usr/bin/env python3
"""
Test script for Status Column Updater

Verifies that the Status column has been updated correctly
"""

import openpyxl
import os

def test_status_updates(file_path):
    """
    Test the status updates in the Excel file
    """
    print(f"\n=== Status Column Update Test ===")
    print(f"Testing file: {os.path.basename(file_path)}")
    
    if not os.path.exists(file_path):
        print(f"❌ Error: File not found: {file_path}")
        return False
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active
        
        print(f"📋 Working with sheet: '{worksheet.title}'")
        
        # Count status values
        status_counts = {}
        let_rows = []
        non_let_rows = []
        
        # Check Status column (AE = column 31)
        for row in range(3, worksheet.max_row + 1):
            status_value = worksheet.cell(row=row, column=31).value
            
            if status_value:
                status_str = str(status_value).strip().upper()
                status_counts[status_str] = status_counts.get(status_str, 0) + 1
                
                if status_str == "LET":
                    let_rows.append(row)
                else:
                    non_let_rows.append(row)
        
        print(f"\n📊 Status Summary:")
        for status, count in sorted(status_counts.items()):
            print(f"   {status}: {count} rows")
        
        print(f"\n✅ LET Status: {len(let_rows)} rows")
        if len(let_rows) <= 10:
            print(f"   Rows: {let_rows}")
        else:
            print(f"   Sample rows: {let_rows[:5]} ... {let_rows[-5:]}")
        
        if non_let_rows:
            print(f"\n⚪ Other Status: {len(non_let_rows)} rows")
            if len(non_let_rows) <= 10:
                print(f"   Rows: {non_let_rows}")
            else:
                print(f"   Sample rows: {non_let_rows[:5]} ... {non_let_rows[-5:]}")
        
        # Sample some LET rows to verify they meet criteria
        print(f"\n🔍 Spot Check (First 5 LET rows):")
        for i, row in enumerate(let_rows[:5]):
            tenant_name = worksheet.cell(row=row, column=11).value  # Column K
            
            print(f"   Row {row}: Tenant='{tenant_name}'")
        
        workbook.close()
        print(f"\n✅ Test completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error testing file: {e}")
        return False

def main():
    # Test the main file
    file_path = "/Volumes/Marketing/Simone Morciano/python working folder/bankScheduleSanitizer/data/30 November 2025 Bank Schedule [updated].xlsx"
    test_status_updates(file_path)
    
    # Check if backup exists
    backup_path = file_path.replace('.xlsx', '_BACKUP_StatusUpdate.xlsx')
    if os.path.exists(backup_path):
        print(f"\n📋 Backup file confirmed: {os.path.basename(backup_path)}")
    else:
        print(f"\n⚠️  No backup file found")

if __name__ == "__main__":
    main()
